import os
import sqlite3
import json
import csv
import io
from datetime import datetime, timedelta
from functools import wraps
from zoneinfo import ZoneInfo

TZ = ZoneInfo('America/Sao_Paulo')


def now_local():
    """Datetime atual em America/Sao_Paulo, sem tzinfo (para armazenar em SQLite)."""
    return datetime.now(TZ).replace(tzinfo=None)


def now_local_str():
    return now_local().strftime('%Y-%m-%d %H:%M:%S')


def _catalog_normalize(nome):
    if not nome:
        return ''
    return ' '.join(nome.strip().upper().split())


def _catalog_find_or_create(db, descricao, unidade=None):
    """Localiza item no catálogo (case-insensitive); cria se novo. Retorna id ou None."""
    nome = _catalog_normalize(descricao)
    if not nome:
        return None
    row = db.execute("SELECT id FROM itens_catalogo WHERE nome = ? COLLATE NOCASE",
                     (nome,)).fetchone()
    if row:
        return row['id'] if isinstance(row, sqlite3.Row) else row[0]
    cur = db.execute(
        "INSERT INTO itens_catalogo (nome, unidade) VALUES (?, ?)",
        (nome, (unidade or '').strip() or None)
    )
    return cur.lastrowid


def _catalog_touch(db, cat_id, when=None):
    """Incrementa contador e atualiza último uso do item."""
    when = when or now_local_str()
    db.execute(
        "UPDATE itens_catalogo SET n_pedidos = n_pedidos + 1, ultimo_uso = ? WHERE id = ?",
        (when, cat_id)
    )


def _catalog_recalc(db):
    """Recalcula n_pedidos e ultimo_uso a partir de requisicoes_itens."""
    db.execute("""
        UPDATE itens_catalogo
        SET n_pedidos = COALESCE((
                SELECT COUNT(*) FROM requisicoes_itens
                WHERE item_catalogo_id = itens_catalogo.id
            ), 0),
            ultimo_uso = (
                SELECT MAX(r.data_solicitacao)
                FROM requisicoes_itens ri
                JOIN requisicoes_compra r ON r.id = ri.requisicao_id
                WHERE ri.item_catalogo_id = itens_catalogo.id
            )
    """)

import bcrypt
import xlrd
import anthropic
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, g, send_file, send_from_directory,
    Response, after_this_request
)
import tempfile
from dotenv import load_dotenv
from embriao_pdf import parse_fivet_pdf
from pecia_pdf import gerar_pdf_estoque_touros, gerar_pdf_estoque_embrioes

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'), override=True)

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-me')

# Sessão persistente: usuário fica logado por 365 dias após cada login,
# mesmo após fechar o navegador. Sai quando clica em "Sair" ou expira.
_IS_DEV = os.getenv('FLASK_ENV') == 'development'
app.permanent_session_lifetime = timedelta(days=365)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=not _IS_DEV,
)

if app.secret_key == 'dev-secret-key-change-me' and not _IS_DEV:
    raise RuntimeError(
        'SECRET_KEY precisa ser configurado em produção. '
        'Gere com: python -c "import secrets; print(secrets.token_hex(32))"'
    )

DATA_DIR = os.getenv('DATA_DIR') or os.path.join(os.path.dirname(__file__), 'data')
DATABASE = os.path.join(DATA_DIR, 'fazenda167.db')
UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER') or os.path.join(os.path.dirname(__file__), 'uploads')
REPORTS_DIR = os.path.join(DATA_DIR, 'relatorios')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)


# ── Database ────────────────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DATABASE)
    with open(os.path.join(os.path.dirname(__file__), 'schema.sql'), encoding='utf-8') as f:
        db.executescript(f.read())

    # Migração: coluna `papel` em usuarios (bancos legados)
    cols = [r[1] for r in db.execute("PRAGMA table_info(usuarios)").fetchall()]
    if 'papel' not in cols:
        db.execute("ALTER TABLE usuarios ADD COLUMN papel TEXT NOT NULL DEFAULT 'usuario'")

    # Migração: coluna `item_catalogo_id` em requisicoes_itens
    cols = [r[1] for r in db.execute("PRAGMA table_info(requisicoes_itens)").fetchall()]
    if 'item_catalogo_id' not in cols:
        db.execute("ALTER TABLE requisicoes_itens ADD COLUMN item_catalogo_id INTEGER REFERENCES itens_catalogo(id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_req_itens_cat ON requisicoes_itens(item_catalogo_id)")

    # Backfill: linka itens existentes ao catálogo e recalcula estatísticas
    db.row_factory = sqlite3.Row
    orfaos = db.execute("""
        SELECT id, descricao, quantidade FROM requisicoes_itens
        WHERE item_catalogo_id IS NULL AND descricao IS NOT NULL AND TRIM(descricao) <> ''
    """).fetchall()
    for it in orfaos:
        cat_id = _catalog_find_or_create(db, it['descricao'], it['quantidade'])
        if cat_id:
            db.execute("UPDATE requisicoes_itens SET item_catalogo_id=? WHERE id=?", (cat_id, it['id']))
    if orfaos:
        _catalog_recalc(db)

    # Bootstrap do usuário master (Dr. Anselmo) a partir das variáveis de ambiente
    admin_email = (os.getenv('ADMIN_EMAIL') or '').strip().lower()
    admin_pass  = os.getenv('ADMIN_PASSWORD')
    if admin_email:
        existing = db.execute(
            "SELECT id FROM usuarios WHERE lower(email)=?", (admin_email,)
        ).fetchone()
        if existing:
            db.execute("UPDATE usuarios SET papel='master' WHERE lower(email)=? AND papel<>'master'",
                       (admin_email,))
        elif admin_pass:
            senha_hash = bcrypt.hashpw(admin_pass.encode(), bcrypt.gensalt()).decode()
            db.execute(
                "INSERT INTO usuarios (nome, email, senha_hash, papel) VALUES (?, ?, ?, 'master')",
                ('Dr. Anselmo', admin_email, senha_hash)
            )

    db.commit()
    db.close()


# ── Auth ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated


def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'erro': 'Não autenticado'}), 401
        return f(*args, **kwargs)
    return decorated


def current_user():
    if 'user_id' not in session:
        return None
    return get_db().execute(
        "SELECT id, nome, email, papel FROM usuarios WHERE id=? AND ativo=1",
        (session['user_id'],)
    ).fetchone()


@app.before_request
def _enforce_active_user():
    """Limpa sessão se o usuário foi desativado — sem isso, sessão de 365 dias
    permitiria acesso de ex-usuários até o cookie expirar."""
    if 'user_id' in session and current_user() is None:
        session.clear()


def is_master(user=None):
    user = user or current_user()
    return bool(user and user['papel'] == 'master')


def master_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        if not is_master():
            return "Acesso restrito ao autorizador (master).", 403
        return f(*args, **kwargs)
    return decorated


def api_master_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'erro': 'Não autenticado'}), 401
        if not is_master():
            return jsonify({'erro': 'Apenas o autorizador (master) pode executar esta ação.'}), 403
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def inject_user_ctx():
    if 'user_id' not in session:
        return {'current_user': None, 'is_master': False}
    user = current_user()
    return {'current_user': user, 'is_master': is_master(user)}


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('password', '').encode()
        user = get_db().execute(
            "SELECT * FROM usuarios WHERE email=? AND ativo=1", (email,)
        ).fetchone()
        if user and bcrypt.checkpw(senha, user['senha_hash'].encode()):
            session.permanent = True
            session['user_id'] = user['id']
            session['user_nome'] = user['nome']
            next_url = request.args.get('next', '/')
            return redirect(next_url)
        flash('E-mail ou senha incorretos.', 'error')
    return render_template('login.html')


@app.route('/instalar')
def instalar():
    """Tutorial pra instalar o sistema como PWA na tela inicial do iPhone.
    Acessível sem login — link enviado por WhatsApp aos usuários."""
    return render_template('instalar.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Helper: encode/decode animal IDs with spaces ───────────────────

def encode_id(animal_id):
    return animal_id.replace(' ', '__') if animal_id else ''


def decode_id(encoded):
    return encoded.replace('__', ' ') if encoded else ''


# ── Helper: get latest rodada ──────────────────────────────────────

def get_ultima_rodada(db):
    return db.execute("SELECT * FROM rodadas ORDER BY id DESC LIMIT 1").fetchone()


def lookup_matriz(doadora):
    """Try to match a doadora ID (free text) against matrizes.animal_id.

    Returns the matched animal_id or None.
    Tries exact match first, then a normalized match (strips spaces/slashes).
    """
    if not doadora:
        return None
    db = get_db()
    m = db.execute(
        "SELECT animal_id FROM matrizes WHERE animal_id=?",
        (doadora,)
    ).fetchone()
    if m:
        return m["animal_id"]
    chave = doadora.replace("/", "").replace(" ", "")
    m = db.execute(
        "SELECT animal_id FROM matrizes "
        "WHERE REPLACE(REPLACE(animal_id,' ',''),'/','')=?",
        (chave,)
    ).fetchone()
    return m["animal_id"] if m else None


# ── Page routes ────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/indices')
@login_required
def indices():
    return render_template('indices.html')


@app.route('/posicao')
@login_required
def posicao():
    return render_template('posicao.html')


@app.route('/matrizes')
@login_required
def matrizes_page():
    return render_template('matrizes.html')


@app.route('/ficha')
@login_required
def ficha():
    return render_template('ficha.html')


@app.route('/touros')
@login_required
def touros():
    return render_template('touros.html')


@app.route('/embrioes')
@login_required
def embrioes_page():
    return render_template('embrioes.html')


@app.route('/embrioes/<int:lote_id>')
@login_required
def embrioes_detalhe_page(lote_id):
    return render_template('embrioes_detalhe.html', lote_id=lote_id)


@app.route('/embrioes/importar')
@login_required
def embrioes_importar_page():
    return render_template('embrioes_importar.html')


@app.route('/embrioes/reconciliar')
@master_required
def embrioes_reconciliar_page():
    return render_template('embrioes_reconciliar.html')


@app.route('/ficha/certidao/<path:animal_id_enc>')
@login_required
def certidao_animal(animal_id_enc):
    return render_template('certidao_animal.html', animal_id_enc=animal_id_enc)


@app.route('/rebanho')
@login_required
def rebanho_page():
    return render_template('rebanho.html')


@app.route('/evolucao')
@login_required
def evolucao_page():
    return render_template('evolucao.html')


@app.route('/safras')
@login_required
def safras_page():
    return render_template('safras.html')


@app.route('/importar')
@login_required
def importar_page():
    return render_template('importar.html')


@app.route('/estoque')
@login_required
def estoque_page():
    return render_template('estoque.html')


@app.route('/estoque/<int:grupo_id>')
@login_required
def estoque_grupo_page(grupo_id):
    return render_template('estoque_grupo.html', grupo_id=grupo_id)


@app.route('/estoque/<int:grupo_id>/catalogo')
@login_required
def catalogo_grupo(grupo_id):
    return render_template('catalogo_touro.html', grupo_id=grupo_id)


# ── API: Dashboard ─────────────────────────────────────────────────

@app.route('/api/dashboard/kpis')
@api_login_required
def api_dashboard_kpis():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'erro': 'Nenhuma rodada importada'}), 404

    rid = rodada['id']

    iciagen_avg = db.execute(
        "SELECT AVG(iciagen) as v FROM avaliacoes WHERE rodada_id=?", (rid,)
    ).fetchone()['v'] or 0

    idesm_avg = db.execute(
        "SELECT AVG(idesm) as v FROM avaliacoes WHERE rodada_id=?", (rid,)
    ).fetchone()['v'] or 0

    ceip_total = db.execute(
        "SELECT COUNT(*) as v FROM matrizes WHERE ceip=1 AND rodada_id=?", (rid,)
    ).fetchone()['v']
    total_mat = db.execute(
        "SELECT COUNT(*) as v FROM matrizes WHERE rodada_id=?", (rid,)
    ).fetchone()['v']

    ipp_avg = db.execute(
        "SELECT AVG(ipp) as v FROM matrizes WHERE ipp IS NOT NULL AND rodada_id=?", (rid,)
    ).fetchone()['v'] or 0

    rmat_avg = db.execute(
        "SELECT AVG(rmat) as v FROM avaliacoes WHERE rmat IS NOT NULL AND rodada_id=?", (rid,)
    ).fetchone()['v'] or 0

    return jsonify({
        'rodada': rodada['nome'],
        'iciagen_avg': round(iciagen_avg, 2),
        'idesm_avg': round(idesm_avg, 2),
        'ceip_total': ceip_total,
        'ceip_pct': round(ceip_total / total_mat * 100, 1) if total_mat else 0,
        'total_matrizes': total_mat,
        'ipp_avg': round(ipp_avg, 1),
        'rmat_avg': round(rmat_avg, 2),
    })


@app.route('/api/dashboard/safras')
@api_login_required
def api_dashboard_safras():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT SUBSTR(m.data_nasc, 7, 4) as safra, AVG(a.iciagen) as avg_icia, COUNT(*) as n
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE m.rodada_id = ? AND m.data_nasc IS NOT NULL
        GROUP BY safra
        ORDER BY safra
    """, (rodada['id'],)).fetchall()

    return jsonify([{'safra': r['safra'], 'avg_icia': round(r['avg_icia'], 2) if r['avg_icia'] is not None else 0, 'n': r['n']} for r in rows])


@app.route('/api/dashboard/composicao')
@api_login_required
def api_dashboard_composicao():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT categoria,
               COUNT(*) as total,
               SUM(genotipada) as genotipadas,
               SUM(precoce) as precoces,
               SUM(ceip) as ceips
        FROM matrizes WHERE rodada_id=?
        GROUP BY categoria ORDER BY categoria
    """, (rodada['id'],)).fetchall()

    return jsonify([dict(r) for r in rows])


@app.route('/api/dashboard/idade')
@api_login_required
def api_dashboard_idade():
    """Distribution of animals by birth year (matrizes + produtos combined)."""
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rid = rodada['id']

    # Matrizes by birth year
    mat_rows = db.execute("""
        SELECT SUBSTR(data_nasc, 7, 4) as ano, COUNT(*) as n
        FROM matrizes WHERE rodada_id=? AND data_nasc IS NOT NULL AND ativo=1
        GROUP BY ano ORDER BY ano
    """, (rid,)).fetchall()

    # Produtos by birth year
    prod_rows = db.execute("""
        SELECT safra_ano as ano, COUNT(*) as n
        FROM produtos WHERE rodada_id=? AND safra_ano IS NOT NULL
        GROUP BY safra_ano ORDER BY safra_ano
    """, (rid,)).fetchall()

    # Merge into one structure
    all_years = {}
    for r in mat_rows:
        if r['ano'] and len(r['ano']) == 4:
            all_years.setdefault(r['ano'], {'matrizes': 0, 'produtos': 0})
            all_years[r['ano']]['matrizes'] = r['n']
    for r in prod_rows:
        if r['ano'] and len(r['ano']) == 4:
            all_years.setdefault(r['ano'], {'matrizes': 0, 'produtos': 0})
            all_years[r['ano']]['produtos'] = r['n']

    result = [{'ano': k, 'matrizes': v['matrizes'], 'produtos': v['produtos']}
              for k, v in sorted(all_years.items())]
    return jsonify(result)


@app.route('/api/dashboard/top10')
@api_login_required
def api_dashboard_top10():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT m.animal_id, m.categoria, a.iciagen, a.deca_icia_g, a.idesm, a.rmat, m.ceip
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE m.rodada_id = ?
        ORDER BY a.iciagen DESC LIMIT 10
    """, (rodada['id'],)).fetchall()

    return jsonify([{**dict(r), 'animal_id_enc': encode_id(r['animal_id'])} for r in rows])


@app.route('/api/dashboard/alertas')
@api_login_required
def api_dashboard_alertas():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'atencao': [], 'destaques': []})

    rows = db.execute("""
        SELECT m.animal_id, a.perc_icia, a.perc_idesm, a.perc_rmat
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE m.rodada_id = ?
    """, (rodada['id'],)).fetchall()

    atencao = []
    destaques = []
    for r in rows:
        if r['perc_icia'] and r['perc_icia'] >= 70:
            atencao.append({'animal_id': r['animal_id'], 'indicador': 'ICIAGen', 'perc': r['perc_icia']})
        if r['perc_icia'] and r['perc_icia'] <= 30:
            destaques.append({'animal_id': r['animal_id'], 'indicador': 'ICIAGen', 'perc': r['perc_icia']})

    atencao.sort(key=lambda x: x['perc'], reverse=True)
    destaques.sort(key=lambda x: x['perc'])

    return jsonify({'atencao': atencao[:10], 'destaques': destaques[:10]})


@app.route('/api/dashboard/dist_deca')
@api_login_required
def api_dashboard_dist_deca():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT deca_icia_g as deca, COUNT(*) as n
        FROM avaliacoes WHERE rodada_id=? AND deca_icia_g IS NOT NULL
        GROUP BY deca_icia_g ORDER BY deca_icia_g
    """, (rodada['id'],)).fetchall()

    return jsonify([dict(r) for r in rows])


# ── API: Matrizes ──────────────────────────────────────────────────

@app.route('/api/matrizes')
@api_login_required
def api_matrizes():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'data': [], 'total': 0})

    rid = rodada['id']
    conditions = ["m.rodada_id = ?"]
    params = [rid]

    q = request.args.get('q', '').strip()
    if q:
        conditions.append("(m.animal_id LIKE ? OR m.touro_pai LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%'])

    ceip = request.args.get('ceip')
    if ceip == '1':
        conditions.append("m.ceip = 1")
    elif ceip == '0':
        conditions.append("m.ceip = 0")

    categ = request.args.get('categ')
    if categ:
        conditions.append("m.categoria = ?")
        params.append(categ)

    deca = request.args.get('deca')
    if deca:
        conditions.append("a.deca_icia_g = ?")
        params.append(int(deca))

    precoce = request.args.get('precoce')
    if precoce == '1':
        conditions.append("m.precoce = 1")
    elif precoce == '0':
        conditions.append("m.precoce = 0")

    iciagen_min = request.args.get('iciagen_min')
    if iciagen_min:
        conditions.append("a.iciagen >= ?")
        params.append(float(iciagen_min))

    idesm_min = request.args.get('idesm_min')
    if idesm_min:
        conditions.append("a.idesm >= ?")
        params.append(float(idesm_min))

    rmat_min = request.args.get('rmat_min')
    if rmat_min:
        conditions.append("a.rmat >= ?")
        params.append(float(rmat_min))

    genotipada = request.args.get('genotipada')
    if genotipada == '1':
        conditions.append("m.genotipada = 1")
    elif genotipada == '0':
        conditions.append("m.genotipada = 0")

    touro_pai = request.args.get('touro_pai')
    if touro_pai:
        conditions.append("m.touro_pai = ?")
        params.append(touro_pai)

    safra = request.args.get('safra')
    if safra:
        conditions.append("SUBSTR(m.data_nasc, 7, 4) = ?")
        params.append(safra)

    ativo_filter = request.args.get('ativo')
    if ativo_filter == '0':
        conditions.append("m.ativo = 0")
    elif ativo_filter != 'todos':
        conditions.append("m.ativo = 1")  # default: only active

    where = " AND ".join(conditions)
    order = request.args.get('order', 'iciagen')
    allowed_orders = {
        'iciagen': 'a.iciagen DESC',
        'idesm': 'a.idesm DESC',
        'rmat': 'a.rmat DESC',
        'ipp': 'm.ipp ASC',
        'pv': 'm.pv DESC',
        'animal_id': 'm.animal_id ASC',
        'categoria': 'm.categoria ASC',
    }
    order_sql = allowed_orders.get(order, 'a.iciagen DESC')

    limit = min(int(request.args.get('limit', 10)), 100)
    offset = int(request.args.get('offset', 0))

    total = db.execute(f"""
        SELECT COUNT(*) as n
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE {where}
    """, params).fetchone()['n']

    rows = db.execute(f"""
        SELECT m.animal_id, m.categoria, m.touro_pai, m.data_nasc, m.ipp, m.pv,
               m.precoce, m.ceip, m.genotipada, m.ativo,
               a.iciagen, a.deca_icia_g, a.deca_icia_f, a.idesm, a.rmat
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE {where}
        ORDER BY {order_sql}
        LIMIT ? OFFSET ?
    """, params + [limit, offset]).fetchall()

    data = []
    for r in rows:
        d = dict(r)
        d['animal_id_enc'] = encode_id(r['animal_id'])
        safra = r['data_nasc'][-4:] if r['data_nasc'] and len(r['data_nasc']) >= 4 else ''
        d['safra'] = safra
        data.append(d)

    return jsonify({'data': data, 'total': total})


@app.route('/api/export/matrizes.csv')
@api_login_required
def api_export_matrizes_csv():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return "Nenhuma rodada", 404

    rows = db.execute("""
        SELECT m.animal_id, m.categoria, m.touro_pai, m.data_nasc, m.ipp, m.iep, m.pv,
               m.precoce, m.ceip, m.genotipada,
               a.iciagen, a.deca_icia_g, a.deca_icia_f, a.idesm, a.deca_idesm_g, a.rmat, a.ifrig
        FROM matrizes m
        JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE m.rodada_id = ?
        ORDER BY a.iciagen DESC
    """, (rodada['id'],)).fetchall()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Animal ID', 'Categoria', 'Touro Pai', 'Nascimento', 'IPP', 'IEP', 'PV',
                     'Precoce', 'CEIP', 'Genotipada', 'ICIAGen', 'Deca G', 'Deca F',
                     'IDESM', 'Deca IDESM', 'RMat', 'IFRIG'])
    for r in rows:
        writer.writerow([r[k] for k in r.keys()])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=matrizes_porto_engenho.csv'}
    )


# ── API: Filter options (for matrizes page) ───────────────────────

@app.route('/api/filtros')
@api_login_required
def api_filtros():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'touros': [], 'safras': []})

    rid = rodada['id']
    touros = db.execute("""
        SELECT DISTINCT touro_pai FROM matrizes
        WHERE rodada_id=? AND touro_pai IS NOT NULL AND touro_pai != ''
        ORDER BY touro_pai
    """, (rid,)).fetchall()

    safras = db.execute("""
        SELECT DISTINCT SUBSTR(data_nasc, 7, 4) as safra FROM matrizes
        WHERE rodada_id=? AND data_nasc IS NOT NULL
        ORDER BY safra
    """, (rid,)).fetchall()

    return jsonify({
        'touros': [r['touro_pai'] for r in touros],
        'safras': [r['safra'] for r in safras if r['safra']],
    })


# ── API: Dropdown ──────────────────────────────────────────────────

@app.route('/api/dropdown')
@api_login_required
def api_dropdown():
    db = get_db()
    result = []

    # Matrizes (all, not just last rodada - so user can find inactive ones too)
    mat_rows = db.execute("""
        SELECT m.animal_id, a.iciagen, m.ceip, m.touro_pai, 'Matriz' as tipo
        FROM matrizes m
        LEFT JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        ORDER BY a.iciagen DESC
    """).fetchall()

    for r in mat_rows:
        result.append({
            'animal_id': r['animal_id'],
            'animal_id_enc': encode_id(r['animal_id']),
            'iciagen': r['iciagen'],
            'ceip': r['ceip'],
            'touro_pai': r['touro_pai'],
            'tipo': 'Matriz',
        })

    # Produtos (latest rodada per product, exclude those already in matrizes)
    prod_rows = db.execute("""
        SELECT p.produto_id as animal_id, p.iciagen, p.ceip, p.touro as touro_pai,
               'Produto' as tipo
        FROM (
            SELECT *, ROW_NUMBER() OVER (PARTITION BY produto_id ORDER BY rodada_id DESC) as rn
            FROM produtos
        ) p
        WHERE p.rn = 1
          AND p.produto_id NOT IN (SELECT animal_id FROM matrizes)
        ORDER BY p.iciagen DESC
    """).fetchall()

    for r in prod_rows:
        result.append({
            'animal_id': r['animal_id'],
            'animal_id_enc': encode_id(r['animal_id']),
            'iciagen': r['iciagen'],
            'ceip': r['ceip'],
            'touro_pai': r['touro_pai'],
            'tipo': 'Produto',
        })

    # Sort all by iciagen desc
    result.sort(key=lambda x: x['iciagen'] or -999, reverse=True)
    return jsonify(result)


# ── API: Animal / Ficha ───────────────────────────────────────────

@app.route('/api/animal/<path:animal_id_enc>')
@api_login_required
def api_animal(animal_id_enc):
    animal_id = decode_id(animal_id_enc)
    db = get_db()

    mat = db.execute("SELECT * FROM matrizes WHERE animal_id=?", (animal_id,)).fetchone()

    # If not found in matrizes, check produtos
    is_produto = False
    if not mat:
        prod = db.execute("""
            SELECT * FROM produtos WHERE produto_id=? ORDER BY rodada_id DESC LIMIT 1
        """, (animal_id,)).fetchone()
        if not prod:
            return jsonify({'erro': 'Animal nao encontrado'}), 404
        # Build a pseudo-matriz dict from produto data
        is_produto = True
        mat = {
            'animal_id': prod['produto_id'], 'data_nasc': prod['data_nasc'],
            'touro_pai': prod['touro'], 'tipo_serv': prod['tipo_serv'],
            'mae_id': prod['mae_id'], 'avo_paterno': None,
            'avo_materno': prod['avo_materno'], 'rebanho': None,
            'genotipada': 0, 'categoria': None, 'precoce': 0,
            'ceip': prod['ceip'], 'np_ceip': 0,
            'score_r': None, 'score_f': None, 'score_a': None, 'score_p': None,
            'rank_cia': None, 'ipp': None, 'iep': None, 'pv': None,
            'desc_ap': None, 'rodada_id': prod['rodada_id'],
            'sexo': prod['sexo'],
        }

    aval = db.execute("""
        SELECT * FROM avaliacoes WHERE animal_id=? ORDER BY rodada_id DESC LIMIT 1
    """, (animal_id,)).fetchone()

    # For produtos, build pseudo avaliacao from product data if no avaliacao
    if not aval and is_produto:
        aval = {
            'iciagen': prod['iciagen'], 'deca_icia_g': prod['deca_iciagen'],
            'deca_icia_f': None, 'perc_icia': None, 'acc_icia': None,
            'idesm': prod['idesm'], 'deca_idesm_g': prod['deca_idesm'],
            'deca_idesm_f': None, 'perc_idesm': None, 'acc_idesm': None,
            'rmat': prod['rmat'], 'deca_rmat': None, 'perc_rmat': None, 'acc_rmat': None,
            'ifrig': prod['ifrig'], 'hgp': None, 'ncaract': None,
            'dep_pn': None, 'dep_gnd': None, 'dep_cd': None, 'dep_pd': None,
            'dep_md': None, 'dep_ud': None, 'dep_gpd': None, 'dep_cs': None,
            'dep_ps': None, 'dep_ms': None, 'dep_us': None,
            'dep_temp': None, 'dep_gns': None, 'dep_pei': None, 'dep_peip': None,
        }

    # Dados da mãe (se estiver no rebanho)
    mae_dados = None
    if mat['mae_id']:
        mae = db.execute("SELECT animal_id FROM matrizes WHERE animal_id=?", (mat['mae_id'],)).fetchone()
        if mae:
            mae_dados = {'animal_id': mae['animal_id'], 'animal_id_enc': encode_id(mae['animal_id'])}

    # Filhos (machos e fêmeas)
    filhos = db.execute("""
        SELECT p.*, r.nome as rodada_nome
        FROM produtos p JOIN rodadas r ON r.id = p.rodada_id
        WHERE p.mae_id = ?
        GROUP BY p.produto_id
        HAVING p.rodada_id = MAX(p.rodada_id)
        ORDER BY p.data_nasc ASC
    """, (animal_id,)).fetchall()

    # Filhas ativas na fazenda (que estão na tabela matrizes)
    filhas_fazenda = db.execute("""
        SELECT m.animal_id, a.iciagen, a.deca_icia_g, m.categoria, m.ceip
        FROM matrizes m
        LEFT JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
        WHERE m.mae_id = ?
        ORDER BY a.iciagen DESC
    """, (animal_id,)).fetchall()

    # Idade calculada
    idade = None
    if mat['data_nasc']:
        try:
            dn = datetime.strptime(mat['data_nasc'], '%d/%m/%Y')
            delta = datetime.now() - dn
            anos = delta.days // 365
            meses = (delta.days % 365) // 30
            idade = f"{anos}a {meses}m"
        except ValueError:
            pass

    # Get rodada info
    rodada_info = None
    if mat['rodada_id']:
        rod = db.execute("SELECT nome FROM rodadas WHERE id=?", (mat['rodada_id'],)).fetchone()
        if rod:
            rodada_info = rod['nome']

    return jsonify({
        'matriz': dict(mat),
        'avaliacao': dict(aval) if aval else None,
        'mae_dados': mae_dados,
        'filhos': [dict(f) for f in filhos],
        'filhas_fazenda': [{**dict(ff), 'animal_id_enc': encode_id(ff['animal_id'])} for ff in filhas_fazenda],
        'idade': idade,
        'animal_id_enc': encode_id(animal_id),
        'rodada_nome': rodada_info,
        'is_produto': is_produto,
    })


@app.route('/api/animal/<path:animal_id_enc>/historico')
@api_login_required
def api_animal_historico(animal_id_enc):
    animal_id = decode_id(animal_id_enc)
    db = get_db()

    rows = db.execute("""
        SELECT a.*, r.nome as rodada_nome
        FROM avaliacoes a JOIN rodadas r ON r.id = a.rodada_id
        WHERE a.animal_id = ?
        ORDER BY r.id
    """, (animal_id,)).fetchall()

    return jsonify([dict(r) for r in rows])


# ── API: Rodadas ───────────────────────────────────────────────────

@app.route('/api/rodadas')
@api_login_required
def api_rodadas():
    db = get_db()
    rows = db.execute("SELECT * FROM rodadas ORDER BY id DESC").fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/rodadas/<int:rid>', methods=['DELETE'])
@api_login_required
def api_delete_rodada(rid):
    db = get_db()
    db.execute("DELETE FROM produtos WHERE rodada_id=?", (rid,))
    db.execute("DELETE FROM avaliacoes WHERE rodada_id=?", (rid,))
    db.execute("DELETE FROM matrizes WHERE rodada_id=?", (rid,))
    db.execute("DELETE FROM rodadas WHERE id=?", (rid,))
    db.commit()
    return jsonify({'ok': True})


# ── API: Importar ──────────────────────────────────────────────────

def safe_float(val):
    if val is None or val == '' or val == '-':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    if val is None or val == '' or val == '-':
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def read_xls_rows(filepath, preferred_sheets=None):
    """Read XLS/XLSX file and return list of dicts (header → value).
    Supports both .xls (xlrd) and .xlsx (openpyxl).
    Auto-detects header row (skips title rows)."""

    if preferred_sheets is None:
        preferred_sheets = ['matrizes', 'geral']

    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        # Find preferred sheet
        ws = None
        for name in preferred_sheets:
            if name in wb.sheetnames:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        # Read all rows as lists
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([v for v in row])
    else:
        # .xls format
        wb = xlrd.open_workbook(filepath)
        sheet = None
        for name in preferred_sheets:
            if name in wb.sheet_names():
                sheet = wb.sheet_by_name(name)
                break
        if sheet is None:
            sheet = wb.sheet_by_index(0)

        all_rows = []
        for i in range(sheet.nrows):
            all_rows.append([sheet.cell_value(i, j) for j in range(sheet.ncols)])

    if not all_rows:
        return [], []

    # Auto-detect header row: find first row with 3+ non-empty distinct values
    header_idx = 0
    for i, row in enumerate(all_rows[:5]):
        non_empty = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(non_empty) >= 3 and len(set(non_empty)) >= 3:
            header_idx = i
            break

    headers = [str(v).strip() if v else '' for v in all_rows[header_idx]]

    rows = []
    for i in range(header_idx + 1, len(all_rows)):
        row_vals = all_rows[i]
        row_data = {}
        for j, h in enumerate(headers):
            val = row_vals[j] if j < len(row_vals) else None
            if h:
                row_data[h] = val
            row_data[f'_col_{j}'] = val
        rows.append(row_data)
    return rows, headers


def format_date(val):
    """Try to format a date value from XLS."""
    if not val:
        return None
    if isinstance(val, float):
        try:
            from xlrd import xldate_as_tuple
            dt = xldate_as_tuple(val, 0)
            return f"{dt[2]:02d}/{dt[1]:02d}/{dt[0]:04d}"
        except Exception:
            return None
    s = str(val).strip()
    if s:
        return s
    return None


@app.route('/api/importar', methods=['POST'])
@api_login_required
def api_importar():
    db = get_db()
    nome_rodada = request.form.get('nome_rodada', '').strip()
    if not nome_rodada:
        return jsonify({'erro': 'Nome da rodada é obrigatório'}), 400

    arquivo_mat = request.files.get('arquivo_mat')
    # Support multiple safra files
    arquivos_saf = request.files.getlist('arquivo_saf')
    # Filter out empty file inputs
    arquivos_saf = [f for f in arquivos_saf if f and f.filename]

    if not arquivo_mat and not arquivos_saf:
        return jsonify({'erro': 'Envie pelo menos um arquivo'}), 400

    saf_names = ', '.join(f.filename for f in arquivos_saf) if arquivos_saf else None

    # Create rodada
    cur = db.execute(
        "INSERT INTO rodadas (nome, arquivo_mat, arquivo_saf) VALUES (?, ?, ?)",
        (nome_rodada,
         arquivo_mat.filename if arquivo_mat else None,
         saf_names)
    )
    rodada_id = cur.lastrowid

    n_matrizes = 0
    n_produtos = 0

    # Import matrizes
    if arquivo_mat:
        mat_path = os.path.join(UPLOAD_FOLDER, arquivo_mat.filename)
        arquivo_mat.save(mat_path)

        rows, headers = read_xls_rows(mat_path)

        # Find indices for duplicate columns (IPP, PV)
        ipp_indices = [i for i, h in enumerate(headers) if h == 'IPP']
        pv_indices = [i for i, h in enumerate(headers) if h == 'PV']
        ipp_col = ipp_indices[-1] if len(ipp_indices) > 1 else (ipp_indices[0] if ipp_indices else None)
        pv_col = pv_indices[-1] if len(pv_indices) > 1 else (pv_indices[0] if pv_indices else None)

        for r in rows:
            animal_id = str(r.get('vaca', '')).strip()
            if not animal_id:
                continue

            try:
                data_nasc = format_date(r.get('dataN'))
                geno_val = r.get('geno', 0)
                genotipada = 1 if geno_val and str(geno_val).strip() not in ('0', '', 'N') else 0
                precoce_val = r.get('precoce', 0)
                precoce = 1 if precoce_val and str(precoce_val).strip() not in ('0', '', 'N') else 0

                # CEIP can be "SIM"/text or number
                ceip_raw = str(r.get('CEIP', '')).strip().upper()
                ceip_flag = 1 if ceip_raw in ('SIM', 'S', '1', 'TRUE') or safe_int(r.get('CEIP')) else 0
                np_ceip = safe_int(r.get('np_CEIP'))

                ipp_val = safe_float(r.get(f'_col_{ipp_col}')) if ipp_col is not None else safe_float(r.get('IPP'))
                pv_val = safe_float(r.get(f'_col_{pv_col}')) if pv_col is not None else safe_float(r.get('PV'))

                # Upsert matrizes
                db.execute("""
                    INSERT INTO matrizes (
                        animal_id, data_nasc, touro_pai, tipo_serv, mae_id,
                        avo_paterno, avo_materno, rebanho, genotipada, categoria,
                        precoce, ceip, np_ceip, score_r, score_f, score_a, score_p,
                        rank_cia, ipp, iep, pv, desc_ap, rodada_id
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(animal_id) DO UPDATE SET
                        data_nasc=excluded.data_nasc, touro_pai=excluded.touro_pai,
                        tipo_serv=excluded.tipo_serv, mae_id=excluded.mae_id,
                        avo_paterno=excluded.avo_paterno, avo_materno=excluded.avo_materno,
                        rebanho=excluded.rebanho, genotipada=excluded.genotipada,
                        categoria=excluded.categoria, precoce=excluded.precoce,
                        ceip=excluded.ceip, np_ceip=excluded.np_ceip,
                        score_r=excluded.score_r, score_f=excluded.score_f,
                        score_a=excluded.score_a, score_p=excluded.score_p,
                        rank_cia=excluded.rank_cia, ipp=excluded.ipp, iep=excluded.iep,
                        pv=excluded.pv, desc_ap=excluded.desc_ap, rodada_id=excluded.rodada_id,
                        updated_at=CURRENT_TIMESTAMP
                """, (
                    animal_id, data_nasc,
                    str(r.get('touro_pai', '')).strip() or None,
                    str(r.get('TS', '')).strip() or None,
                    str(r.get('mae', '')).strip() or None,
                    str(r.get('avo_paterno', '')).strip() or None,
                    str(r.get('avo_materno', '')).strip() or None,
                    str(r.get('rebanho', '')).strip() or None,
                    genotipada,
                    str(r.get('categ', '')).strip() or None,
                    precoce, ceip_flag, np_ceip,
                    safe_float(r.get('R')), safe_float(r.get('F')),
                    safe_float(r.get('A')), safe_float(r.get('P')),
                    safe_int(r.get('rank')),
                    ipp_val,
                    safe_float(r.get('IEP')),
                    pv_val,
                    str(r.get('DESC_AP', '')).strip() or None,
                    rodada_id
                ))

                # Insert avaliação
                perc_icia = safe_float(r.get('perc_ICiaGen'))
                if perc_icia is not None and perc_icia <= 1:
                    perc_icia = round(perc_icia * 100, 2)

                perc_idesm = safe_float(r.get('perc_IDESM'))
                if perc_idesm is not None and perc_idesm <= 1:
                    perc_idesm = round(perc_idesm * 100, 2)

                perc_rmat = safe_float(r.get('perc_RMat'))
                if perc_rmat is not None and perc_rmat <= 1:
                    perc_rmat = round(perc_rmat * 100, 2)

                db.execute("""
                    INSERT OR REPLACE INTO avaliacoes (
                        animal_id, rodada_id,
                        iciagen, deca_icia_g, deca_icia_f, perc_icia, acc_icia,
                        idesm, deca_idesm_g, deca_idesm_f, perc_idesm, acc_idesm,
                        rmat, deca_rmat, perc_rmat, acc_rmat,
                        ifrig, hgp, ncaract,
                        dep_pn, dep_gnd, dep_cd, dep_pd, dep_md, dep_ud,
                        dep_gpd, dep_cs, dep_ps, dep_ms, dep_us,
                        dep_temp, dep_gns, dep_pei, dep_peip
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    animal_id, rodada_id,
                    safe_float(r.get('ICiaGen')),
                    safe_int(r.get('decaG_ICiaGen')),
                    safe_int(r.get('decaF_ICiaGen')),
                    perc_icia,
                    safe_float(r.get('acc_ICiaGen')),
                    safe_float(r.get('IDESM')),
                    safe_int(r.get('decaG_IDESM')),
                    safe_int(r.get('decaF_IDESM')),
                    perc_idesm,
                    safe_float(r.get('acc_IDESM')),
                    safe_float(r.get('RMat')),
                    safe_int(r.get('decaG_RMat')),
                    perc_rmat,
                    safe_float(r.get('acc_RMat')),
                    safe_float(r.get('IFRIG')),
                    safe_float(r.get('HGP')),
                    safe_int(r.get('ncaract')),
                    safe_float(r.get('PN')),
                    safe_float(r.get('GND')),
                    safe_float(r.get('CD')),
                    safe_float(r.get('PD')),
                    safe_float(r.get('MD')),
                    safe_float(r.get('UD')),
                    safe_float(r.get('GPD')),
                    safe_float(r.get('CS')),
                    safe_float(r.get('PS')),
                    safe_float(r.get('MS')),
                    safe_float(r.get('US')),
                    safe_float(r.get('TEMP')),
                    safe_float(r.get('GNS')),
                    safe_float(r.get('PEi')),
                    safe_float(r.get('PEip')),
                ))

                n_matrizes += 1
            except Exception as e:
                app.logger.warning(f"Erro ao importar matriz {animal_id}: {e}")
                continue

    # Import produtos (safra) — supports multiple files
    for arquivo_saf in arquivos_saf:
        saf_path = os.path.join(UPLOAD_FOLDER, arquivo_saf.filename)
        arquivo_saf.save(saf_path)

        rows, headers = read_xls_rows(saf_path)

        for r in rows:
            produto_id = str(r.get('produto', '')).strip()
            if not produto_id:
                continue

            try:
                # Handle boolean-like fields (can be "SIM", "S", 1, etc.)
                pai_dna_raw = str(r.get('pai_dna', '')).strip().upper()
                pai_dna = 1 if pai_dna_raw in ('SIM', 'S', '1', 'TRUE') or safe_int(r.get('pai_dna')) else 0
                ceip_raw = str(r.get('CEIP', '')).strip().upper()
                ceip_prod = 1 if ceip_raw in ('SIM', 'S', '1', 'TRUE') or safe_int(r.get('CEIP')) else 0

                data_nasc_prod = format_date(r.get('dataN'))
                safra_ano = data_nasc_prod[-4:] if data_nasc_prod and len(data_nasc_prod) >= 4 else None

                db.execute("""
                    INSERT OR IGNORE INTO produtos (
                        produto_id, rodada_id, mae_id, touro, tipo_serv, avo_materno,
                        data_nasc, sexo, pn, peso_desm, idade_desm, peso_sob, idade_sob,
                        gnd_aj, gpd_aj, iciagen, deca_iciagen, idesm, deca_idesm,
                        rmat, ifrig, conect_desm, pai_dna, ceip, safra_ano
                    ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    produto_id, rodada_id,
                    str(r.get('vaca', '')).strip() or None,
                    str(r.get('touro', '')).strip() or None,
                    str(r.get('TS', '')).strip() or None,
                    str(r.get('avo_materno', '')).strip() or None,
                    data_nasc_prod,
                    str(r.get('sexo', '')).strip() or None,
                    safe_float(r.get('PN')),
                    safe_float(r.get('peso_DESM')),
                    safe_int(r.get('idade_DESM')),
                    safe_float(r.get('peso_SOB')),
                    safe_int(r.get('idade_SOB')),
                    safe_float(r.get('aj_GND')),
                    safe_float(r.get('aj_GPD')),
                    safe_float(r.get('ICiaGen')),
                    safe_int(r.get('decaG_ICiaGen')),
                    safe_float(r.get('IDESM')),
                    safe_int(r.get('decaG_IDESM')),
                    safe_float(r.get('RMat')),
                    safe_float(r.get('IFRIG')),
                    str(r.get('conect_DESM', '')).strip() or None,
                    pai_dna,
                    ceip_prod,
                    safra_ano,
                ))
                n_produtos += 1
            except Exception as e:
                app.logger.warning(f"Erro ao importar produto {produto_id}: {e}")
                continue

    # Mark animals not in this import as inactive (only if matrizes were imported)
    n_inativadas = 0
    if arquivo_mat and n_matrizes > 0:
        cur_inativ = db.execute(
            "UPDATE matrizes SET ativo=0 WHERE rodada_id != ? AND ativo=1",
            (rodada_id,)
        )
        n_inativadas = cur_inativ.rowcount
        # Ensure imported ones are active
        db.execute("UPDATE matrizes SET ativo=1 WHERE rodada_id=?", (rodada_id,))

    # Update rodada counts
    db.execute("UPDATE rodadas SET n_matrizes=?, n_produtos=? WHERE id=?",
               (n_matrizes, n_produtos, rodada_id))
    db.commit()

    return jsonify({
        'ok': True,
        'rodada_id': rodada_id,
        'n_matrizes': n_matrizes,
        'n_produtos': n_produtos,
        'n_inativadas': n_inativadas,
    })


# ── API: Touros ────────────────────────────────────────────────────

@app.route('/api/touros/contribuicao')
@api_login_required
def api_touros_contribuicao():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT touro_pai as touro, COUNT(*) as n
        FROM matrizes WHERE rodada_id=? AND touro_pai IS NOT NULL AND touro_pai != ''
        GROUP BY touro_pai ORDER BY n DESC LIMIT 15
    """, (rodada['id'],)).fetchall()

    total = db.execute(
        "SELECT COUNT(*) as v FROM matrizes WHERE rodada_id=? AND touro_pai IS NOT NULL",
        (rodada['id'],)
    ).fetchone()['v']

    result = []
    acum = 0
    for r in rows:
        pct = round(r['n'] / total * 100, 2) if total else 0
        acum += pct
        result.append({'touro': r['touro'], 'n': r['n'], 'pct': pct, 'acum': round(acum, 2)})

    return jsonify(result)


@app.route('/api/touros/iciagen_safra')
@api_login_required
def api_touros_iciagen_safra():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify([])

    rows = db.execute("""
        SELECT p.touro, SUBSTR(p.data_nasc, 7, 4) as safra, AVG(p.iciagen) as avg_icia, COUNT(*) as n
        FROM produtos p
        WHERE p.rodada_id=? AND p.touro IS NOT NULL AND p.data_nasc IS NOT NULL
        GROUP BY p.touro, safra
        ORDER BY avg_icia DESC
    """, (rodada['id'],)).fetchall()

    return jsonify([dict(r) for r in rows])


# ── API: Posição vs CIA ────────────────────────────────────────────

@app.route('/api/posicao')
@api_login_required
def api_posicao():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({})

    rid = rodada['id']

    # Average percentiles
    percs = db.execute("""
        SELECT
            AVG(perc_icia) as perc_icia,
            AVG(perc_idesm) as perc_idesm,
            AVG(perc_rmat) as perc_rmat
        FROM avaliacoes WHERE rodada_id=?
    """, (rid,)).fetchone()

    # IPP position
    ipp = db.execute("""
        SELECT AVG(ipp) as v FROM matrizes WHERE rodada_id=? AND ipp IS NOT NULL
    """, (rid,)).fetchone()['v']

    # CEIP %
    ceip_pct = db.execute("""
        SELECT CAST(SUM(ceip) AS REAL) / COUNT(*) * 100 as v
        FROM matrizes WHERE rodada_id=?
    """, (rid,)).fetchone()['v'] or 0

    # Precoces %
    precoce_pct = db.execute("""
        SELECT CAST(SUM(precoce) AS REAL) / COUNT(*) * 100 as v
        FROM matrizes WHERE rodada_id=?
    """, (rid,)).fetchone()['v'] or 0

    # Genotipadas %
    geno_pct = db.execute("""
        SELECT CAST(SUM(genotipada) AS REAL) / COUNT(*) * 100 as v
        FROM matrizes WHERE rodada_id=?
    """, (rid,)).fetchone()['v'] or 0

    # ICIAGen by category
    cats = db.execute("""
        SELECT m.categoria, AVG(a.iciagen) as avg_icia
        FROM matrizes m JOIN avaliacoes a ON a.animal_id=m.animal_id AND a.rodada_id=m.rodada_id
        WHERE m.rodada_id=?
        GROUP BY m.categoria ORDER BY m.categoria
    """, (rid,)).fetchall()

    return jsonify({
        'perc_icia': round(percs['perc_icia'] or 0, 1),
        'perc_idesm': round(percs['perc_idesm'] or 0, 1),
        'perc_rmat': round(percs['perc_rmat'] or 0, 1),
        'ipp_avg': round(ipp or 0, 1),
        'ceip_pct': round(ceip_pct, 1),
        'precoce_pct': round(precoce_pct, 1),
        'geno_pct': round(geno_pct, 1),
        'categorias': [dict(c) for c in cats],
    })


# ── API: Estoque de Touros (por grupo/safra) ──────────────────────

def _get_val(r, *keys):
    """Get first non-empty value from multiple possible column names."""
    for k in keys:
        v = r.get(k)
        if v is not None and str(v).strip():
            return v
    return None


@app.route('/api/estoque/grupos')
@api_login_required
def api_estoque_grupos():
    db = get_db()
    grupos = db.execute("SELECT * FROM estoque_grupos ORDER BY id DESC").fetchall()
    result = []
    for g in grupos:
        stats = db.execute("""
            SELECT COUNT(*) as total,
                   SUM(CASE WHEN vendido=0 THEN 1 ELSE 0 END) as disp,
                   SUM(CASE WHEN vendido=1 THEN 1 ELSE 0 END) as vend,
                   AVG(CASE WHEN vendido=0 THEN iciagen END) as avg_icia,
                   AVG(CASE WHEN vendido=1 THEN valor_venda END) as avg_venda
            FROM estoque_touros WHERE grupo_id=?
        """, (g['id'],)).fetchone()
        result.append({
            **dict(g),
            'total': stats['total'],
            'disp': stats['disp'] or 0,
            'vend': stats['vend'] or 0,
            'avg_icia': round(stats['avg_icia'], 2) if stats['avg_icia'] else 0,
            'avg_venda': round(stats['avg_venda'], 2) if stats['avg_venda'] else 0,
        })
    return jsonify(result)


@app.route('/api/estoque/grupos', methods=['POST'])
@api_login_required
def api_estoque_criar_grupo():
    db = get_db()
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()
    if not nome:
        return jsonify({'erro': 'Nome é obrigatório'}), 400
    try:
        cur = db.execute("INSERT INTO estoque_grupos (nome) VALUES (?)", (nome,))
        db.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Já existe um grupo com esse nome'}), 400


@app.route('/api/estoque/grupos/<int:gid>', methods=['DELETE'])
@api_login_required
def api_estoque_del_grupo(gid):
    db = get_db()
    db.execute("DELETE FROM estoque_touros WHERE grupo_id=?", (gid,))
    db.execute("DELETE FROM estoque_grupos WHERE id=?", (gid,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estoque/grupos/<int:gid>')
@api_login_required
def api_estoque_grupo(gid):
    db = get_db()
    grupo = db.execute("SELECT * FROM estoque_grupos WHERE id=?", (gid,)).fetchone()
    if not grupo:
        return jsonify({'erro': 'Grupo não encontrado'}), 404

    todos = request.args.get('todos', '0')
    q = request.args.get('q', '').strip()

    conditions = ["grupo_id = ?"]
    params = [gid]

    if todos != '1':
        conditions.append("vendido = 0")
    if q:
        conditions.append("(brinco LIKE ? OR pai LIKE ? OR avo_paterno LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%'])

    where = " AND ".join(conditions)

    rows = db.execute(f"""
        SELECT * FROM estoque_touros WHERE {where} ORDER BY iciagen DESC
    """, params).fetchall()

    # KPIs for this group
    kpis = db.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN vendido=0 THEN 1 ELSE 0 END) as disp,
               SUM(CASE WHEN vendido=1 THEN 1 ELSE 0 END) as vend,
               AVG(CASE WHEN vendido=0 THEN iciagen END) as avg_icia,
               AVG(CASE WHEN vendido=0 THEN idesm END) as avg_idesm,
               AVG(CASE WHEN vendido=0 THEN rmat END) as avg_rmat,
               AVG(CASE WHEN vendido=1 THEN valor_venda END) as avg_venda,
               SUM(CASE WHEN vendido=1 THEN valor_venda ELSE 0 END) as total_venda
        FROM estoque_touros WHERE grupo_id=?
    """, (gid,)).fetchone()

    return jsonify({
        'grupo': dict(grupo),
        'data': [dict(r) for r in rows],
        'kpis': {
            'total': kpis['total'],
            'disp': kpis['disp'] or 0,
            'vend': kpis['vend'] or 0,
            'avg_icia': round(kpis['avg_icia'], 2) if kpis['avg_icia'] else 0,
            'avg_idesm': round(kpis['avg_idesm'], 2) if kpis['avg_idesm'] else 0,
            'avg_rmat': round(kpis['avg_rmat'], 2) if kpis['avg_rmat'] else 0,
            'avg_venda': round(kpis['avg_venda'], 2) if kpis['avg_venda'] else 0,
            'total_venda': round(kpis['total_venda'] or 0, 2),
        }
    })


@app.route('/api/estoque/grupos/<int:gid>/importar', methods=['POST'])
@api_login_required
def api_estoque_importar_grupo(gid):
    db = get_db()
    grupo = db.execute("SELECT id FROM estoque_grupos WHERE id=?", (gid,)).fetchone()
    if not grupo:
        return jsonify({'erro': 'Grupo não encontrado'}), 404

    arquivo = request.files.get('arquivo')
    if not arquivo or not arquivo.filename:
        return jsonify({'erro': 'Envie um arquivo XLS/XLSX'}), 400

    filepath = os.path.join(UPLOAD_FOLDER, arquivo.filename)
    arquivo.save(filepath)
    rows, headers = read_xls_rows(filepath)
    n = 0

    app.logger.info(f"Estoque import headers: {headers}")

    for r in rows:
        brinco = str(_get_val(r, 'BRINCO', 'brinco', 'Brinco', 'produto', 'ID', 'id') or '').strip()
        if not brinco:
            continue

        idesm_val = _get_val(r, 'IDESM', 'idesm', 'Idesm', 'IDESM R')

        try:
            db.execute("""
                INSERT INTO estoque_touros (brinco, grupo_id, data_nasc, pai, avo_paterno,
                    avo_materno, idesm, iciagen, rmat, ifrig, peso, data_pesagem)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(brinco, grupo_id) DO UPDATE SET
                    data_nasc=excluded.data_nasc, pai=excluded.pai,
                    avo_paterno=excluded.avo_paterno, avo_materno=excluded.avo_materno,
                    idesm=excluded.idesm, iciagen=excluded.iciagen,
                    rmat=excluded.rmat, ifrig=excluded.ifrig,
                    peso=excluded.peso, data_pesagem=excluded.data_pesagem
            """, (
                brinco, gid,
                format_date(_get_val(r, 'data de Nascimento', 'dataN', 'Data de Nascimento', 'data_nasc', 'Nascimento')),
                str(_get_val(r, 'Pai', 'pai', 'PAI', 'touro_pai', 'touro') or '').strip() or None,
                str(_get_val(r, 'Avô Paterno', 'avo_paterno', 'Avo Paterno', 'AVO_PATERNO') or '').strip() or None,
                str(_get_val(r, 'Avô Materno', 'avo_materno', 'Avo Materno', 'AVO_MATERNO') or '').strip() or None,
                safe_float(idesm_val),
                safe_float(_get_val(r, 'ICiaGen', 'iciagen', 'ICIAGen', 'ICIAGEN')),
                safe_float(_get_val(r, 'RMat', 'rmat', 'RMAT')),
                safe_float(_get_val(r, 'IFRIG', 'ifrig', 'Ifrig')),
                safe_float(_get_val(r, 'Peso', 'peso', 'PESO', 'PV')),
                format_date(_get_val(r, 'data_pesagem', 'Data Pesagem', 'data pesagem')),
            ))
            n += 1
        except Exception as e:
            app.logger.warning(f"Erro ao importar touro {brinco}: {e}")
            continue

    db.commit()
    return jsonify({'ok': True, 'n_touros': n})


@app.route('/api/estoque/<int:touro_id>', methods=['PUT'])
@api_login_required
def api_estoque_editar(touro_id):
    db = get_db()
    data = request.get_json() or {}
    fields = []
    params = []
    allowed = ['brinco', 'data_nasc', 'pai', 'avo_paterno', 'avo_materno',
               'idesm', 'iciagen', 'rmat', 'ifrig', 'peso', 'data_pesagem', 'obs']
    for f in allowed:
        if f in data:
            fields.append(f"{f}=?")
            val = data[f]
            if f in ('idesm', 'iciagen', 'rmat', 'ifrig', 'peso'):
                val = safe_float(val)
            elif val == '':
                val = None
            params.append(val)
    if not fields:
        return jsonify({'erro': 'Nenhum campo para atualizar'}), 400
    params.append(touro_id)
    db.execute(f"UPDATE estoque_touros SET {', '.join(fields)} WHERE id=?", params)
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estoque/<int:touro_id>/vender', methods=['POST'])
@api_login_required
def api_estoque_vender(touro_id):
    db = get_db()
    data = request.get_json() or {}
    valor = safe_float(data.get('valor'))
    db.execute(
        "UPDATE estoque_touros SET vendido=1, data_venda=DATE('now'), valor_venda=? WHERE id=?",
        (valor, touro_id)
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estoque/<int:touro_id>/devolver', methods=['POST'])
@api_login_required
def api_estoque_devolver(touro_id):
    db = get_db()
    db.execute(
        "UPDATE estoque_touros SET vendido=0, data_venda=NULL, valor_venda=NULL WHERE id=?",
        (touro_id,)
    )
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estoque/<int:touro_id>', methods=['DELETE'])
@api_login_required
def api_estoque_delete(touro_id):
    db = get_db()
    db.execute("DELETE FROM estoque_touros WHERE id=?", (touro_id,))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/estoque/grupos/<int:gid>/renomear', methods=['POST'])
@api_login_required
def api_estoque_renomear(gid):
    db = get_db()
    data = request.get_json() or {}
    nome = data.get('nome', '').strip()
    if not nome:
        return jsonify({'erro': 'Nome obrigatorio'}), 400
    try:
        db.execute("UPDATE estoque_grupos SET nome=? WHERE id=?", (nome, gid))
        db.commit()
        return jsonify({'ok': True})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Ja existe um grupo com esse nome'}), 400


# ── API: Rebanho (todos os animais, última rodada) ────────────────

@app.route('/api/rebanho')
@api_login_required
def api_rebanho():
    db = get_db()
    ultima = get_ultima_rodada(db)
    ultima_id = ultima['id'] if ultima else 0

    q = request.args.get('q', '').strip()
    ativo = request.args.get('ativo', '')
    tipo = request.args.get('tipo', '')  # 'matriz', 'produto', or '' for all
    categ = request.args.get('categ', '')
    sexo = request.args.get('sexo', '')
    order = request.args.get('order', 'iciagen')
    limit = min(int(request.args.get('limit', 20)), 200)
    offset = int(request.args.get('offset', 0))

    # Build a unified view using a CTE
    # Matrizes: use their rodada_id and avaliacao data
    # Produtos: use their rodada_id and inline genetic data
    # "ativo" = animal is in the latest rodada
    base_sql = f"""
        SELECT animal_id, tipo, categoria, pai, data_nasc, sexo,
               iciagen, deca_icia, idesm, rmat, ipp, pv,
               ceip, precoce, genotipada,
               rodada_id, rodada_nome,
               CASE WHEN rodada_id = {ultima_id} THEN 1 ELSE 0 END as ativo
        FROM (
            SELECT m.animal_id, 'Matriz' as tipo, m.categoria, m.touro_pai as pai,
                   m.data_nasc, NULL as sexo,
                   a.iciagen, a.deca_icia_g as deca_icia, a.idesm, a.rmat,
                   m.ipp, m.pv, m.ceip, m.precoce, m.genotipada,
                   m.rodada_id, r.nome as rodada_nome
            FROM matrizes m
            LEFT JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
            LEFT JOIN rodadas r ON r.id = m.rodada_id

            UNION ALL

            SELECT p.produto_id as animal_id, 'Produto' as tipo, NULL as categoria,
                   p.touro as pai, p.data_nasc, p.sexo,
                   p.iciagen, p.deca_iciagen as deca_icia, p.idesm, p.rmat,
                   NULL as ipp, NULL as pv, p.ceip, 0 as precoce, 0 as genotipada,
                   p.rodada_id, r2.nome as rodada_nome
            FROM (
                SELECT *, ROW_NUMBER() OVER (PARTITION BY produto_id ORDER BY rodada_id DESC) as rn
                FROM produtos
            ) p
            LEFT JOIN rodadas r2 ON r2.id = p.rodada_id
            WHERE p.rn = 1
        ) reb
    """

    conditions = []
    params = []

    if q:
        conditions.append("(animal_id LIKE ? OR pai LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%'])
    if ativo == '1':
        conditions.append("ativo = 1")
    elif ativo == '0':
        conditions.append("ativo = 0")
    if tipo == 'matriz':
        conditions.append("tipo = 'Matriz'")
    elif tipo == 'produto':
        conditions.append("tipo = 'Produto'")
    if categ:
        conditions.append("categoria = ?")
        params.append(categ)
    if sexo:
        conditions.append("sexo = ?")
        params.append(sexo)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

    allowed_orders = {
        'iciagen': 'iciagen DESC', 'idesm': 'idesm DESC',
        'rmat': 'rmat DESC', 'ipp': 'ipp ASC',
        'pv': 'pv DESC', 'animal_id': 'animal_id ASC',
        'categoria': 'categoria ASC', 'rodada': 'rodada_id DESC',
        'tipo': 'tipo ASC',
    }
    order_sql = allowed_orders.get(order, 'iciagen DESC')

    count_sql = f"SELECT COUNT(*) as n FROM ({base_sql}) sub {where}"
    total = db.execute(count_sql, params).fetchone()['n']

    data_sql = f"SELECT * FROM ({base_sql}) sub {where} ORDER BY {order_sql} LIMIT ? OFFSET ?"
    rows = db.execute(data_sql, params + [limit, offset]).fetchall()

    # KPIs
    kpi_sql = f"""SELECT COUNT(*) as total,
        SUM(CASE WHEN ativo=1 THEN 1 ELSE 0 END) as ativos,
        SUM(CASE WHEN ativo=0 THEN 1 ELSE 0 END) as inativos,
        SUM(CASE WHEN tipo='Matriz' THEN 1 ELSE 0 END) as n_matrizes,
        SUM(CASE WHEN tipo='Produto' THEN 1 ELSE 0 END) as n_produtos
    FROM ({base_sql}) sub"""
    kpis = db.execute(kpi_sql).fetchone()

    data = []
    for r in rows:
        d = dict(r)
        d['animal_id_enc'] = encode_id(r['animal_id'])
        data.append(d)

    return jsonify({
        'data': data,
        'total': total,
        'kpis': {
            'total': kpis['total'],
            'ativos': kpis['ativos'] or 0,
            'inativos': kpis['inativos'] or 0,
            'n_matrizes': kpis['n_matrizes'] or 0,
            'n_produtos': kpis['n_produtos'] or 0,
        }
    })


# ── API: Evolução do Rebanho ───────────────────────────────────────

@app.route('/api/evolucao')
@api_login_required
def api_evolucao():
    db = get_db()
    rodadas = db.execute("SELECT * FROM rodadas ORDER BY id ASC").fetchall()

    result = []
    for rod in rodadas:
        rid = rod['id']
        # Matrizes stats
        mat_stats = db.execute("""
            SELECT COUNT(*) as n,
                   AVG(a.iciagen) as avg_icia, AVG(a.idesm) as avg_idesm,
                   AVG(a.rmat) as avg_rmat, AVG(m.ipp) as avg_ipp,
                   SUM(m.ceip) as n_ceip, SUM(m.precoce) as n_precoce,
                   SUM(m.genotipada) as n_geno
            FROM matrizes m
            JOIN avaliacoes a ON a.animal_id = m.animal_id AND a.rodada_id = m.rodada_id
            WHERE m.rodada_id = ?
        """, (rid,)).fetchone()

        # Produtos stats per safra year
        prod_safras = db.execute("""
            SELECT safra_ano, COUNT(*) as n,
                   AVG(iciagen) as avg_icia, AVG(idesm) as avg_idesm,
                   SUM(CASE WHEN sexo='M' THEN 1 ELSE 0 END) as machos,
                   SUM(CASE WHEN sexo='F' THEN 1 ELSE 0 END) as femeas
            FROM produtos WHERE rodada_id=? AND safra_ano IS NOT NULL
            GROUP BY safra_ano ORDER BY safra_ano
        """, (rid,)).fetchall()

        n_produtos = db.execute(
            "SELECT COUNT(*) as v FROM produtos WHERE rodada_id=?", (rid,)
        ).fetchone()['v']

        result.append({
            'rodada': dict(rod),
            'matrizes': {
                'n': mat_stats['n'] or 0,
                'avg_icia': round(mat_stats['avg_icia'], 2) if mat_stats['avg_icia'] else 0,
                'avg_idesm': round(mat_stats['avg_idesm'], 2) if mat_stats['avg_idesm'] else 0,
                'avg_rmat': round(mat_stats['avg_rmat'], 2) if mat_stats['avg_rmat'] else 0,
                'avg_ipp': round(mat_stats['avg_ipp'], 1) if mat_stats['avg_ipp'] else 0,
                'n_ceip': mat_stats['n_ceip'] or 0,
                'n_precoce': mat_stats['n_precoce'] or 0,
                'n_geno': mat_stats['n_geno'] or 0,
            },
            'produtos': {
                'n': n_produtos,
                'safras': [{
                    'ano': s['safra_ano'],
                    'n': s['n'],
                    'machos': s['machos'] or 0,
                    'femeas': s['femeas'] or 0,
                    'avg_icia': round(s['avg_icia'], 2) if s['avg_icia'] else 0,
                    'avg_idesm': round(s['avg_idesm'], 2) if s['avg_idesm'] else 0,
                } for s in prod_safras],
            },
        })

    return jsonify(result)


# ── API: Safras (produtos por ano/rodada) ─────────────────────────

@app.route('/api/safras')
@api_login_required
def api_safras():
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'rodada': None, 'anos': []})

    rid = rodada['id']
    anos = db.execute("""
        SELECT safra_ano, COUNT(*) as n,
               SUM(CASE WHEN sexo='M' THEN 1 ELSE 0 END) as machos,
               SUM(CASE WHEN sexo='F' THEN 1 ELSE 0 END) as femeas,
               AVG(iciagen) as avg_icia, AVG(idesm) as avg_idesm
        FROM produtos WHERE rodada_id=? AND safra_ano IS NOT NULL
        GROUP BY safra_ano ORDER BY safra_ano
    """, (rid,)).fetchall()

    return jsonify({
        'rodada': dict(rodada),
        'anos': [{**dict(a),
                  'avg_icia': round(a['avg_icia'], 2) if a['avg_icia'] else 0,
                  'avg_idesm': round(a['avg_idesm'], 2) if a['avg_idesm'] else 0,
                  } for a in anos]
    })


@app.route('/api/safras/<ano>')
@api_login_required
def api_safra_detalhe(ano):
    db = get_db()
    rodada = get_ultima_rodada(db)
    if not rodada:
        return jsonify({'data': [], 'total': 0})

    rid = rodada['id']
    q = request.args.get('q', '').strip()
    sexo = request.args.get('sexo', '')
    order = request.args.get('order', 'iciagen')
    limit = min(int(request.args.get('limit', 20)), 100)
    offset = int(request.args.get('offset', 0))

    conditions = ["rodada_id=?", "safra_ano=?"]
    params = [rid, ano]

    if q:
        conditions.append("(produto_id LIKE ? OR touro LIKE ? OR mae_id LIKE ?)")
        params.extend([f'%{q}%', f'%{q}%', f'%{q}%'])
    if sexo:
        conditions.append("sexo=?")
        params.append(sexo)

    touro_filter = request.args.get('touro', '').strip()
    if touro_filter:
        conditions.append("touro=?")
        params.append(touro_filter)

    ceip_filter = request.args.get('ceip')
    if ceip_filter == '1':
        conditions.append("ceip=1")

    where = " AND ".join(conditions)
    allowed_orders = {
        'iciagen': 'iciagen DESC', 'idesm': 'idesm DESC',
        'rmat': 'rmat DESC', 'produto_id': 'produto_id ASC',
        'data_nasc': 'data_nasc ASC',
    }
    order_sql = allowed_orders.get(order, 'iciagen DESC')

    total = db.execute(f"SELECT COUNT(*) as n FROM produtos WHERE {where}", params).fetchone()['n']
    rows = db.execute(f"""
        SELECT * FROM produtos WHERE {where} ORDER BY {order_sql} LIMIT ? OFFSET ?
    """, params + [limit, offset]).fetchall()

    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'rodada': dict(rodada)})


# ── PecIA: página ──────────────────────────────────────────────────

@app.route('/pecia')
@login_required
def page_pecia():
    return render_template('pecia.html')


# ── Requisições de Compra ──────────────────────────────────────────

def _req_proximo_numero(db):
    ano = datetime.now().year
    row = db.execute(
        "SELECT COUNT(*) AS n FROM requisicoes_compra WHERE numero LIKE ?",
        (f'REQ-{ano}-%',)
    ).fetchone()
    return f"REQ-{ano}-{(row['n'] + 1):04d}"


def _req_log(db, req_id, acao, detalhes=None):
    user = current_user()
    db.execute(
        """INSERT INTO requisicoes_historico
           (requisicao_id, usuario_id, usuario_nome, acao, detalhes, created_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (req_id,
         user['id'] if user else None,
         user['nome'] if user else None,
         acao, detalhes, now_local_str())
    )


def _req_fetch(db, req_id):
    req = db.execute(
        "SELECT * FROM requisicoes_compra WHERE id=?", (req_id,)
    ).fetchone()
    if not req:
        return None, None, None
    itens = db.execute(
        "SELECT * FROM requisicoes_itens WHERE requisicao_id=? ORDER BY ordem, id",
        (req_id,)
    ).fetchall()
    hist = db.execute(
        "SELECT * FROM requisicoes_historico WHERE requisicao_id=? ORDER BY created_at DESC, id DESC",
        (req_id,)
    ).fetchall()
    return req, itens, hist


@app.route('/requisicoes')
@login_required
def requisicoes_page():
    return render_template('requisicoes.html')


@app.route('/requisicoes/nova')
@login_required
def requisicao_nova_page():
    return render_template('requisicao_nova.html')


@app.route('/requisicoes/<int:req_id>')
@login_required
def requisicao_detalhe_page(req_id):
    return render_template('requisicao_detalhe.html', req_id=req_id)


@app.route('/api/requisicoes', methods=['GET'])
@api_login_required
def api_requisicoes_list():
    db = get_db()
    user = current_user()
    status = request.args.get('status', '').strip()
    escopo = request.args.get('escopo', '').strip()  # 'meus' | 'todos'

    where = []
    params = []
    if status in ('pendente', 'aprovada', 'rejeitada', 'cancelada'):
        where.append("r.status = ?")
        params.append(status)

    # Usuário comum só vê suas próprias requisições; master vê tudo (ou 'meus' opcional)
    if not is_master(user) or escopo == 'meus':
        where.append("r.solicitante_id = ?")
        params.append(user['id'])

    sql = """
        SELECT r.*,
               (SELECT COUNT(*) FROM requisicoes_itens WHERE requisicao_id=r.id) AS n_itens
        FROM requisicoes_compra r
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY CASE r.status WHEN 'pendente' THEN 0 ELSE 1 END, r.id DESC"

    rows = db.execute(sql, params).fetchall()
    return jsonify({
        'is_master': is_master(user),
        'requisicoes': [dict(r) for r in rows],
    })


@app.route('/api/requisicoes', methods=['POST'])
@api_login_required
def api_requisicoes_criar():
    db = get_db()
    user = current_user()
    data = request.get_json(silent=True) or {}

    responsavel = (data.get('responsavel') or 'Fazenda Porto do Engenho').strip()
    funcionario_retirada = (data.get('funcionario_retirada') or '').strip()
    fornecedor = (data.get('fornecedor') or '').strip()
    observacoes = (data.get('observacoes') or '').strip()
    itens_raw = data.get('itens') or []

    itens = []
    for i, it in enumerate(itens_raw):
        desc = (it.get('descricao') or '').strip() if isinstance(it, dict) else ''
        qtd = (it.get('quantidade') or '').strip() if isinstance(it, dict) else ''
        if desc:
            itens.append((i + 1, desc, qtd))

    if not funcionario_retirada:
        return jsonify({'erro': 'Informe o funcionário que fará a retirada.'}), 400
    if not fornecedor:
        return jsonify({'erro': 'Informe o fornecedor.'}), 400
    if not itens:
        return jsonify({'erro': 'Adicione ao menos um item.'}), 400

    numero = _req_proximo_numero(db)
    ts = now_local_str()
    cur = db.execute(
        """INSERT INTO requisicoes_compra
           (numero, data_solicitacao, solicitante_id, solicitante_nome, responsavel,
            funcionario_retirada, fornecedor, observacoes, status, created_at, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pendente', ?, ?)""",
        (numero, ts, user['id'], user['nome'], responsavel,
         funcionario_retirada, fornecedor, observacoes, ts, ts)
    )
    req_id = cur.lastrowid
    for ordem, desc, qtd in itens:
        cat_id = _catalog_find_or_create(db, desc, qtd)
        if cat_id:
            _catalog_touch(db, cat_id, ts)
        db.execute(
            "INSERT INTO requisicoes_itens (requisicao_id, ordem, descricao, quantidade, item_catalogo_id) VALUES (?, ?, ?, ?, ?)",
            (req_id, ordem, desc, qtd, cat_id)
        )
    _req_log(db, req_id, 'criada', f'{len(itens)} item(ns)')
    db.commit()
    return jsonify({'id': req_id, 'numero': numero}), 201


@app.route('/api/requisicoes/<int:req_id>', methods=['GET'])
@api_login_required
def api_requisicao_get(req_id):
    db = get_db()
    user = current_user()
    req, itens, hist = _req_fetch(db, req_id)
    if not req:
        return jsonify({'erro': 'Requisição não encontrada.'}), 404
    if not is_master(user) and req['solicitante_id'] != user['id']:
        return jsonify({'erro': 'Sem permissão.'}), 403
    return jsonify({
        'is_master': is_master(user),
        'requisicao': dict(req),
        'itens': [dict(r) for r in itens],
        'historico': [dict(r) for r in hist],
    })


@app.route('/api/requisicoes/<int:req_id>', methods=['PUT'])
@api_login_required
def api_requisicao_editar(req_id):
    """Edita requisição — permitido apenas quando pendente."""
    db = get_db()
    user = current_user()
    req = db.execute(
        "SELECT solicitante_id, status FROM requisicoes_compra WHERE id=?", (req_id,)
    ).fetchone()
    if not req:
        return jsonify({'erro': 'Requisição não encontrada.'}), 404
    if req['solicitante_id'] != user['id'] and not is_master(user):
        return jsonify({'erro': 'Sem permissão.'}), 403
    if req['status'] != 'pendente':
        return jsonify({'erro': 'Só é possível editar requisições pendentes.'}), 400

    data = request.get_json(silent=True) or {}
    responsavel = (data.get('responsavel') or 'Fazenda Porto do Engenho').strip()
    funcionario_retirada = (data.get('funcionario_retirada') or '').strip()
    fornecedor = (data.get('fornecedor') or '').strip()
    observacoes = (data.get('observacoes') or '').strip()
    itens_raw = data.get('itens') or []

    itens = []
    for i, it in enumerate(itens_raw):
        desc = (it.get('descricao') or '').strip() if isinstance(it, dict) else ''
        qtd = (it.get('quantidade') or '').strip() if isinstance(it, dict) else ''
        if desc:
            itens.append((i + 1, desc, qtd))

    if not funcionario_retirada:
        return jsonify({'erro': 'Informe o funcionário que fará a retirada.'}), 400
    if not fornecedor:
        return jsonify({'erro': 'Informe o fornecedor.'}), 400
    if not itens:
        return jsonify({'erro': 'Adicione ao menos um item.'}), 400

    ts = now_local_str()
    db.execute("""
        UPDATE requisicoes_compra
        SET responsavel=?, funcionario_retirada=?, fornecedor=?, observacoes=?, updated_at=?
        WHERE id=?
    """, (responsavel, funcionario_retirada, fornecedor, observacoes, ts, req_id))

    # Substitui itens (edição completa)
    db.execute("DELETE FROM requisicoes_itens WHERE requisicao_id=?", (req_id,))
    for ordem, desc, qtd in itens:
        cat_id = _catalog_find_or_create(db, desc, qtd)
        db.execute(
            "INSERT INTO requisicoes_itens (requisicao_id, ordem, descricao, quantidade, item_catalogo_id) VALUES (?, ?, ?, ?, ?)",
            (req_id, ordem, desc, qtd, cat_id)
        )
    # Recalcula contadores do catálogo (edição não deve inflar n_pedidos)
    _catalog_recalc(db)
    _req_log(db, req_id, 'editada', f'{len(itens)} item(ns)')
    db.commit()
    return jsonify({'ok': True})


@app.route('/requisicoes/<int:req_id>/editar')
@login_required
def requisicao_editar_page(req_id):
    return render_template('requisicao_editar.html', req_id=req_id)


def _aprovar_uma(db, req_id, assinatura):
    req = db.execute(
        "SELECT id, status, numero FROM requisicoes_compra WHERE id=?", (req_id,)
    ).fetchone()
    if not req:
        return False, 'não encontrada'
    if req['status'] != 'pendente':
        return False, f"não está pendente (status: {req['status']})"
    user = current_user()
    ts = now_local_str()
    db.execute(
        """UPDATE requisicoes_compra
           SET status='aprovada', aprovador_id=?, aprovador_nome=?, assinatura=?,
               data_decisao=?, updated_at=?
           WHERE id=?""",
        (user['id'], user['nome'], assinatura, ts, ts, req_id)
    )
    _req_log(db, req_id, 'aprovada', f'Assinado por {user["nome"]}')
    return True, req['numero']


@app.route('/api/requisicoes/<int:req_id>/aprovar', methods=['POST'])
@api_master_required
def api_requisicao_aprovar(req_id):
    db = get_db()
    # Assinatura sempre é o nome do master logado — não pode ser sobrescrito pelo cliente
    assinatura = current_user()['nome']
    ok, info = _aprovar_uma(db, req_id, assinatura)
    if not ok:
        return jsonify({'erro': f'Requisição {info}.'}), 400
    db.commit()
    return jsonify({'ok': True, 'numero': info})


@app.route('/api/requisicoes/aprovar-lote', methods=['POST'])
@api_master_required
def api_requisicao_aprovar_lote():
    db = get_db()
    data = request.get_json(silent=True) or {}
    ids = data.get('ids') or []
    if not ids or not isinstance(ids, list):
        return jsonify({'erro': 'Selecione ao menos uma requisição.'}), 400
    # Assinatura sempre é o nome do master logado
    assinatura = current_user()['nome']

    aprovadas, falhas = [], []
    for rid in ids:
        try:
            rid_int = int(rid)
        except (TypeError, ValueError):
            falhas.append({'id': rid, 'motivo': 'id inválido'})
            continue
        ok, info = _aprovar_uma(db, rid_int, assinatura)
        if ok:
            aprovadas.append({'id': rid_int, 'numero': info})
        else:
            falhas.append({'id': rid_int, 'motivo': info})
    db.commit()
    return jsonify({'aprovadas': aprovadas, 'falhas': falhas})


@app.route('/api/requisicoes/<int:req_id>/rejeitar', methods=['POST'])
@api_master_required
def api_requisicao_rejeitar(req_id):
    db = get_db()
    data = request.get_json(silent=True) or {}
    motivo = (data.get('motivo') or '').strip()
    if not motivo:
        return jsonify({'erro': 'Informe o motivo da rejeição.'}), 400
    req = db.execute(
        "SELECT status FROM requisicoes_compra WHERE id=?", (req_id,)
    ).fetchone()
    if not req:
        return jsonify({'erro': 'Requisição não encontrada.'}), 404
    if req['status'] != 'pendente':
        return jsonify({'erro': f'Requisição não está pendente (status: {req["status"]}).'}), 400
    user = current_user()
    ts = now_local_str()
    db.execute(
        """UPDATE requisicoes_compra
           SET status='rejeitada', aprovador_id=?, aprovador_nome=?,
               motivo_rejeicao=?, data_decisao=?, updated_at=?
           WHERE id=?""",
        (user['id'], user['nome'], motivo, ts, ts, req_id)
    )
    _req_log(db, req_id, 'rejeitada', motivo)
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/requisicoes/<int:req_id>/cancelar', methods=['POST'])
@api_login_required
def api_requisicao_cancelar(req_id):
    db = get_db()
    user = current_user()
    data = request.get_json(silent=True) or {}
    motivo = (data.get('motivo') or '').strip()

    req = db.execute(
        "SELECT solicitante_id, status FROM requisicoes_compra WHERE id=?", (req_id,)
    ).fetchone()
    if not req:
        return jsonify({'erro': 'Requisição não encontrada.'}), 404
    if req['solicitante_id'] != user['id'] and not is_master(user):
        return jsonify({'erro': 'Sem permissão.'}), 403
    if req['status'] in ('cancelada', 'rejeitada'):
        return jsonify({'erro': f'Requisição já está {req["status"]}.'}), 400
    if req['status'] == 'aprovada' and not motivo:
        return jsonify({'erro': 'Para cancelar uma requisição já autorizada, informe o motivo.'}), 400

    db.execute(
        """UPDATE requisicoes_compra
           SET status='cancelada', updated_at=?
           WHERE id=?""",
        (now_local_str(), req_id)
    )
    detalhes = motivo or None
    if req['status'] == 'aprovada' and motivo:
        detalhes = f'Cancelada após autorização. Motivo: {motivo}'
    _req_log(db, req_id, 'cancelada', detalhes)
    db.commit()
    return jsonify({'ok': True})


@app.route('/requisicoes/<int:req_id>/pdf')
@login_required
def requisicao_pdf(req_id):
    db = get_db()
    user = current_user()
    req, itens, _ = _req_fetch(db, req_id)
    if not req:
        return "Requisição não encontrada.", 404
    if not is_master(user) and req['solicitante_id'] != user['id']:
        return "Sem permissão.", 403
    if req['status'] != 'aprovada':
        return "A autorização em PDF só é emitida após aprovação.", 400

    pdf_bytes = _gerar_pdf_requisicao(req, itens)
    _req_log(db, req_id, 'pdf_gerado')
    db.commit()
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'inline; filename=autorizacao_{req["numero"]}.pdf'
        }
    )


def _fmt_dt(s):
    if not s:
        return ''
    try:
        if isinstance(s, str):
            dt = datetime.fromisoformat(s.replace('Z', '').split('.')[0])
        else:
            dt = s
        return dt.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return str(s)


def _gerar_pdf_requisicao(req, itens):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    margin = 2 * cm
    red = colors.HexColor('#D01B20')
    black = colors.HexColor('#1A1A1A')
    gray = colors.HexColor('#666666')

    # Cabeçalho com logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.jpg')
    try:
        if os.path.exists(logo_path):
            c.drawImage(logo_path, margin, h - margin - 2.2*cm, width=3.2*cm, height=2.2*cm,
                        preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(margin + 3.8*cm, h - margin - 0.8*cm, 'FAZENDA PORTO DO ENGENHO')
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin + 3.8*cm, h - margin - 1.4*cm, 'AUTORIZAÇÃO DE COMPRA')
    c.setFont('Helvetica', 9)
    c.setFillColor(gray)
    c.drawString(margin + 3.8*cm, h - margin - 2.0*cm, f'Nº {req["numero"]}')

    # Linha vermelha
    c.setStrokeColor(red)
    c.setLineWidth(2)
    y = h - margin - 2.6*cm
    c.line(margin, y, w - margin, y)

    # Metadados
    y -= 0.8*cm
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 10)
    c.drawString(margin, y, 'Data da solicitação:')
    c.setFont('Helvetica', 10)
    c.drawString(margin + 4.0*cm, y, _fmt_dt(req['data_solicitacao']))

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Responsável pela solicitação:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 5.6*cm, y, req['responsavel'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Funcionário que fará a retirada:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 5.9*cm, y, req['funcionario_retirada'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Fornecedor:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 2.4*cm, y, req['fornecedor'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Solicitante (login):')
    c.setFont('Helvetica', 10);       c.drawString(margin + 3.6*cm, y, req['solicitante_nome'] or '')

    # Itens
    y -= 0.9*cm
    c.setFont('Helvetica-Bold', 11); c.drawString(margin, y, 'Materiais autorizados:')
    y -= 0.4*cm

    data = [['#', 'Quantidade', 'Descrição']]
    for it in itens:
        data.append([str(it['ordem']), it['quantidade'] or '', it['descricao'] or ''])
    tbl = Table(data, colWidths=[1.0*cm, 3.2*cm, w - 2*margin - 4.2*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), red),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',(0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
    ]))
    tw, th = tbl.wrap(w - 2*margin, y)
    tbl.drawOn(c, margin, y - th)
    y = y - th - 0.5*cm

    # Observações
    if req['observacoes']:
        c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Observações:')
        y -= 0.4*cm
        c.setFont('Helvetica', 9)
        for linha in req['observacoes'].splitlines() or [req['observacoes']]:
            c.drawString(margin, y, linha[:110])
            y -= 0.4*cm

    # Bloco de autorização
    y -= 0.4*cm
    c.setStrokeColor(red); c.setLineWidth(1.2)
    box_h = 3.8*cm
    c.rect(margin, y - box_h, w - 2*margin, box_h, stroke=1, fill=0)
    c.setFillColor(red)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin + 0.4*cm, y - 0.6*cm, 'AUTORIZAÇÃO')
    c.setFillColor(black)
    c.setFont('Helvetica', 10)
    c.drawString(margin + 0.4*cm, y - 1.2*cm,
                 'Autorizo a retirada dos materiais acima listados junto ao fornecedor indicado.')
    c.setFont('Helvetica-Bold', 10)
    c.drawString(margin + 0.4*cm, y - 2.0*cm, 'Autorizado por:')
    c.setFont('Helvetica', 10)
    c.drawString(margin + 3.5*cm, y - 2.0*cm, req['aprovador_nome'] or '')
    c.setFont('Helvetica-Bold', 10)
    c.drawString(margin + 0.4*cm, y - 2.6*cm, 'Data/Hora:')
    c.setFont('Helvetica', 10)
    c.drawString(margin + 2.4*cm, y - 2.6*cm, _fmt_dt(req['data_decisao']))
    c.setFont('Helvetica-Bold', 10)
    c.drawString(margin + 0.4*cm, y - 3.2*cm, 'Assinatura digital:')
    c.setFont('Helvetica-Oblique', 11)
    c.setFillColor(red)
    c.drawString(margin + 3.7*cm, y - 3.2*cm, req['assinatura'] or '')
    c.setFillColor(black)

    # Rodapé
    c.setFont('Helvetica', 7); c.setFillColor(gray)
    c.drawString(margin, margin - 0.6*cm,
                 f'Documento gerado eletronicamente · Nº {req["numero"]} · '
                 f'{now_local().strftime("%d/%m/%Y %H:%M")}')
    c.drawRightString(w - margin, margin - 0.6*cm,
                      '(27) 3225-8853 · fazendaportodoengenho@gmail.com')

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.read()


# ── PDF: requisição individual com espaço para assinar manualmente ─

@app.route('/requisicoes/<int:req_id>/imprimir.pdf')
@login_required
def requisicao_imprimir_pdf(req_id):
    """PDF de uma requisição (qualquer status) com bloco de assinatura em branco
    para assinar à mão. Distinto da rota /pdf, que só gera o comprovante oficial
    pós-aprovação."""
    db = get_db()
    user = current_user()
    req, itens, _ = _req_fetch(db, req_id)
    if not req:
        return "Requisição não encontrada.", 404
    if not is_master(user) and req['solicitante_id'] != user['id']:
        return "Sem permissão.", 403

    pdf_bytes = _gerar_pdf_imprimir(req, itens)
    stamp = now_local().strftime('%Y%m%d_%H%M')
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'inline; filename=requisicao_{req["numero"]}_{stamp}.pdf',
            'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
            'Pragma': 'no-cache',
            'Expires': '0',
        }
    )


def _gerar_pdf_imprimir(req, itens):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    margin = 2 * cm
    red = colors.HexColor('#D01B20')
    black = colors.HexColor('#1A1A1A')
    gray = colors.HexColor('#666666')

    status_pt = {
        'pendente':   'PENDENTE',
        'aprovada':   'AUTORIZADA',
        'rejeitada':  'REJEITADA',
        'cancelada':  'CANCELADA',
    }.get(req['status'], req['status'].upper())

    # Cabeçalho com logo
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.jpg')
    try:
        if os.path.exists(logo_path):
            c.drawImage(logo_path, margin, h - margin - 2.2*cm, width=3.2*cm, height=2.2*cm,
                        preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 14)
    c.drawString(margin + 3.8*cm, h - margin - 0.8*cm, 'FAZENDA PORTO DO ENGENHO')
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin + 3.8*cm, h - margin - 1.4*cm, 'REQUISIÇÃO DE COMPRA')
    c.setFont('Helvetica', 9)
    c.setFillColor(gray)
    c.drawString(margin + 3.8*cm, h - margin - 2.0*cm,
                 f'Nº {req["numero"]} · Status atual: {status_pt}')

    c.setStrokeColor(red); c.setLineWidth(2)
    y = h - margin - 2.6*cm
    c.line(margin, y, w - margin, y)

    # Metadados
    y -= 0.8*cm
    c.setFillColor(black)
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Data da solicitação:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 4.0*cm, y, _fmt_dt(req['data_solicitacao']))

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Solicitante:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 2.5*cm, y, req['solicitante_nome'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Responsável pela solicitação:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 5.6*cm, y, req['responsavel'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Funcionário que fará a retirada:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 5.9*cm, y, req['funcionario_retirada'] or '')

    y -= 0.55*cm
    c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Fornecedor:')
    c.setFont('Helvetica', 10);       c.drawString(margin + 2.4*cm, y, req['fornecedor'] or '')

    # Itens
    y -= 0.9*cm
    c.setFont('Helvetica-Bold', 11); c.drawString(margin, y, 'Materiais solicitados:')
    y -= 0.4*cm

    data = [['#', 'Quantidade', 'Descrição']]
    for it in itens:
        data.append([str(it['ordem']), it['quantidade'] or '', it['descricao'] or ''])
    if not itens:
        data.append(['-', '-', '(sem itens)'])
    tbl = Table(data, colWidths=[1.0*cm, 3.2*cm, w - 2*margin - 4.2*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), red),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7F7F7')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING',(0, 0), (-1, -1), 6),
        ('TOPPADDING',  (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
    ]))
    tw, th = tbl.wrap(w - 2*margin, y)
    tbl.drawOn(c, margin, y - th)
    y = y - th - 0.5*cm

    # Observações
    if req['observacoes']:
        c.setFont('Helvetica-Bold', 10); c.drawString(margin, y, 'Observações:')
        y -= 0.4*cm
        c.setFont('Helvetica', 9)
        for linha in req['observacoes'].splitlines() or [req['observacoes']]:
            c.drawString(margin, y, linha[:110])
            y -= 0.4*cm

    # Bloco de assinatura manual
    y -= 0.5*cm
    c.setStrokeColor(red); c.setLineWidth(1.2)
    box_h = 4.6*cm
    c.rect(margin, y - box_h, w - 2*margin, box_h, stroke=1, fill=0)

    c.setFillColor(red); c.setFont('Helvetica-Bold', 11)
    c.drawString(margin + 0.4*cm, y - 0.6*cm, 'AUTORIZAÇÃO MANUAL')

    c.setFillColor(black); c.setFont('Helvetica', 9)
    c.drawString(margin + 0.4*cm, y - 1.2*cm,
                 'Marque a opção desejada e assine para autorizar ou rejeitar:')

    # Checkboxes APROVAR / REJEITAR
    c.setLineWidth(0.8)
    c.rect(margin + 0.4*cm, y - 2.05*cm, 0.5*cm, 0.5*cm, stroke=1, fill=0)
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin + 1.1*cm, y - 1.9*cm, 'APROVAR')

    c.rect(margin + 5.5*cm, y - 2.05*cm, 0.5*cm, 0.5*cm, stroke=1, fill=0)
    c.drawString(margin + 6.2*cm, y - 1.9*cm, 'REJEITAR')

    # Linha de assinatura + data
    c.setFont('Helvetica', 9)
    c.drawString(margin + 0.4*cm, y - 3.2*cm, 'Assinatura:')
    c.line(margin + 2.4*cm, y - 3.2*cm, margin + 12.0*cm, y - 3.2*cm)
    c.drawString(margin + 12.6*cm, y - 3.2*cm, 'Data:')
    c.line(margin + 13.8*cm, y - 3.2*cm, w - margin - 0.4*cm, y - 3.2*cm)

    # Nome em letra de forma
    c.drawString(margin + 0.4*cm, y - 4.1*cm, 'Nome (letra de forma):')
    c.line(margin + 4.8*cm, y - 4.1*cm, w - margin - 0.4*cm, y - 4.1*cm)

    # Rodapé
    c.setFont('Helvetica', 7); c.setFillColor(gray)
    c.drawString(margin, margin - 0.6*cm,
                 f'Documento gerado eletronicamente · Nº {req["numero"]} · '
                 f'{now_local().strftime("%d/%m/%Y %H:%M")}')
    c.drawRightString(w - margin, margin - 0.6*cm,
                      '(27) 3225-8853 · fazendaportodoengenho@gmail.com')

    c.showPage()
    c.save()
    buf.seek(0)
    return buf.read()


# ── PDF: lista de pendentes para assinatura física ─────────────────

@app.route('/requisicoes/pendentes.pdf')
@login_required
def requisicoes_pendentes_pdf():
    """Gera um PDF com todas as requisições pendentes — para assinar em papel."""
    db = get_db()
    user = current_user()

    where = ["r.status = 'pendente'"]
    params = []
    if not is_master(user):
        where.append("r.solicitante_id = ?")
        params.append(user['id'])
    where_sql = " AND ".join(where)

    reqs = db.execute(f"""
        SELECT r.id, r.numero, r.data_solicitacao, r.solicitante_nome,
               r.responsavel, r.funcionario_retirada, r.fornecedor, r.observacoes
        FROM requisicoes_compra r
        WHERE {where_sql}
        ORDER BY r.id ASC
    """, params).fetchall()

    itens_por_req = {}
    for req in reqs:
        rows = db.execute(
            "SELECT ordem, descricao, quantidade FROM requisicoes_itens WHERE requisicao_id=? ORDER BY ordem",
            (req['id'],)
        ).fetchall()
        itens_por_req[req['id']] = rows

    pdf_bytes = _gerar_pdf_pendentes(reqs, itens_por_req, master=is_master(user))
    stamp = now_local().strftime('%Y%m%d_%H%M')
    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={
            'Content-Disposition': f'inline; filename=requisicoes_pendentes_{stamp}.pdf',
            'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
            'Pragma': 'no-cache',
            'Expires': '0',
        }
    )


def _gerar_pdf_pendentes(reqs, itens_por_req, master=False):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak, KeepTogether
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf = io.BytesIO()
    margin = 1.7 * cm
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=1.5*cm,
        title='Requisições Pendentes'
    )
    W = A4[0] - 2 * margin

    red   = colors.HexColor('#D01B20')
    black = colors.HexColor('#1A1A1A')
    gray  = colors.HexColor('#666666')
    lgray = colors.HexColor('#CCCCCC')

    s = getSampleStyleSheet()
    st_title     = ParagraphStyle('t',  parent=s['Title'],   fontName='Helvetica-Bold', fontSize=14, textColor=black, leading=16, spaceAfter=2)
    st_subtitle  = ParagraphStyle('st', parent=s['Normal'],  fontName='Helvetica-Bold', fontSize=11, textColor=red,   leading=13, spaceAfter=0)
    st_meta      = ParagraphStyle('m',  parent=s['Normal'],  fontSize=8,  textColor=gray, leading=10)
    st_req_num   = ParagraphStyle('rn', parent=s['Normal'],  fontName='Helvetica-Bold', fontSize=12, textColor=black, leading=14)
    st_req_sub   = ParagraphStyle('rs', parent=s['Normal'],  fontSize=8.5, textColor=gray, leading=10, spaceAfter=4)
    st_field     = ParagraphStyle('f',  parent=s['Normal'],  fontSize=9, leading=11, textColor=black)
    st_obs       = ParagraphStyle('o',  parent=s['Normal'],  fontSize=8.5, leading=11, textColor=gray, leftIndent=0)
    st_footer    = ParagraphStyle('fo', parent=s['Normal'],  fontSize=7,   textColor=gray, alignment=TA_CENTER)

    story = []

    # ── Cabeçalho ──
    logo_path = os.path.join(os.path.dirname(__file__), 'static', 'logo.jpg')
    header_items = []
    if os.path.exists(logo_path):
        header_items.append([Image(logo_path, width=2.6*cm, height=1.8*cm),
                             [Paragraph('FAZENDA PORTO DO ENGENHO', st_title),
                              Paragraph('Requisições pendentes de autorização', st_subtitle),
                              Paragraph(f'Emitido em {now_local().strftime("%d/%m/%Y %H:%M")} · '
                                        f'{len(reqs)} requisição(ões) aguardando', st_meta)]])
        t = Table(header_items, colWidths=[3.2*cm, W - 3.2*cm])
        t.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ]))
        story.append(t)
    else:
        story.append(Paragraph('FAZENDA PORTO DO ENGENHO', st_title))
        story.append(Paragraph('Requisições pendentes de autorização', st_subtitle))
        story.append(Paragraph(f'Emitido em {now_local().strftime("%d/%m/%Y %H:%M")}', st_meta))

    story.append(Spacer(1, 0.2*cm))
    linha = Table([['']], colWidths=[W])
    linha.setStyle(TableStyle([('LINEBELOW', (0,0), (-1,-1), 1.5, red)]))
    story.append(linha)
    story.append(Spacer(1, 0.4*cm))

    # ── Sem pendentes ──
    if not reqs:
        story.append(Spacer(1, 2*cm))
        story.append(Paragraph(
            '<para align="center"><font size="11" color="#666">Não há requisições pendentes de autorização no momento.</font></para>',
            s['Normal']
        ))
        doc.build(story)
        buf.seek(0)
        return buf.read()

    # ── Cada requisição em bloco (KeepTogether pra não quebrar meio) ──
    for idx, req in enumerate(reqs):
        block = []
        header_row = Table(
            [[Paragraph(f"<b>{req['numero']}</b>", st_req_num),
              Paragraph(f"Solicitada em {_fmt_dt(req['data_solicitacao'])} por <b>{req['solicitante_nome'] or ''}</b>", st_req_sub)]],
            colWidths=[4.0*cm, W - 4.0*cm]
        )
        header_row.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F4F4F4')),
            ('BOX', (0, 0), (-1, -1), 0.5, lgray),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        block.append(header_row)

        fields = [
            [Paragraph('<b>Responsável:</b>', st_field),
             Paragraph(req['responsavel'] or '-', st_field),
             Paragraph('<b>Fornecedor:</b>', st_field),
             Paragraph(req['fornecedor'] or '-', st_field)],
            [Paragraph('<b>Retirada:</b>', st_field),
             Paragraph(req['funcionario_retirada'] or '-', st_field),
             Paragraph('', st_field),
             Paragraph('', st_field)],
        ]
        fields_t = Table(fields, colWidths=[2.6*cm, (W-2*2.6*cm)/2, 2.6*cm, (W-2*2.6*cm)/2])
        fields_t.setStyle(TableStyle([
            ('LINEBELOW', (0, 0), (-1, -1), 0.25, lgray),
            ('BOX', (0, 0), (-1, -1), 0.5, lgray),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        block.append(fields_t)

        # Tabela de itens
        itens = itens_por_req.get(req['id'], [])
        itens_data = [['#', 'Qtd', 'Descrição do item']]
        for it in itens:
            itens_data.append([str(it['ordem']), it['quantidade'] or '', it['descricao'] or ''])
        if not itens:
            itens_data.append(['-', '-', '(sem itens)'])
        itens_t = Table(itens_data, colWidths=[0.8*cm, 2.2*cm, W - 0.8*cm - 2.2*cm])
        itens_t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), red),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 8.5),
            ('VALIGN',     (0, 0), (-1, -1), 'TOP'),
            ('GRID',       (0, 0), (-1, -1), 0.25, lgray),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFAFA')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',(0, 0), (-1, -1), 5),
            ('TOPPADDING',  (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ]))
        block.append(itens_t)

        if req['observacoes']:
            block.append(Spacer(1, 0.12*cm))
            block.append(Paragraph(f"<b>Obs.:</b> {req['observacoes']}", st_obs))

        # Área de decisão/assinatura
        block.append(Spacer(1, 0.15*cm))
        decisao = Table(
            [[Paragraph('<font size="10"><b>(&nbsp;&nbsp;&nbsp;)</b> APROVAR</font>', st_field),
              Paragraph('<font size="10"><b>(&nbsp;&nbsp;&nbsp;)</b> REJEITAR</font>', st_field),
              Paragraph('<font size="9">Assinatura: ____________________________________</font>', st_field)]],
            colWidths=[2.6*cm, 2.6*cm, W - 5.2*cm]
        )
        decisao.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFFCF2')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E6A817')),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        block.append(decisao)
        block.append(Spacer(1, 0.5*cm))

        story.append(KeepTogether(block))

    # ── Rodapé simples ──
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f'Documento gerado eletronicamente · Fazenda Porto do Engenho · '
        f'{now_local().strftime("%d/%m/%Y %H:%M")}',
        st_footer
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Catálogo de itens ──────────────────────────────────────────────

@app.route('/api/itens-catalogo')
@api_login_required
def api_itens_catalogo():
    db = get_db()
    q = (request.args.get('q') or '').strip()
    if q:
        rows = db.execute("""
            SELECT id, nome, unidade, categoria, n_pedidos, ultimo_uso
            FROM itens_catalogo
            WHERE ativo = 1 AND nome LIKE ? COLLATE NOCASE
            ORDER BY n_pedidos DESC, nome
            LIMIT 50
        """, (f'%{q}%',)).fetchall()
    else:
        rows = db.execute("""
            SELECT id, nome, unidade, categoria, n_pedidos, ultimo_uso
            FROM itens_catalogo
            WHERE ativo = 1
            ORDER BY n_pedidos DESC, nome
            LIMIT 500
        """).fetchall()
    return jsonify({'itens': [dict(r) for r in rows]})


# ── Relatório de autorizações ──────────────────────────────────────

@app.route('/requisicoes/relatorio')
@login_required
def requisicoes_relatorio_page():
    return render_template('requisicoes_relatorio.html')


def _relatorio_query(db, user, args):
    data_ini = (args.get('data_ini') or '').strip()
    data_fim = (args.get('data_fim') or '').strip()
    item_id  = args.get('item_id')
    fornecedor = (args.get('fornecedor') or '').strip()

    where = ["r.status = 'aprovada'"]
    params = []
    if not is_master(user):
        where.append("r.solicitante_id = ?")
        params.append(user['id'])
    if data_ini:
        where.append("DATE(r.data_decisao) >= DATE(?)")
        params.append(data_ini)
    if data_fim:
        where.append("DATE(r.data_decisao) <= DATE(?)")
        params.append(data_fim)
    if item_id and str(item_id).isdigit():
        where.append("ri.item_catalogo_id = ?")
        params.append(int(item_id))
    if fornecedor:
        where.append("r.fornecedor LIKE ? COLLATE NOCASE")
        params.append(f'%{fornecedor}%')
    where_sql = " AND ".join(where)

    linhas = db.execute(f"""
        SELECT r.id AS req_id, r.numero, r.data_decisao, r.data_solicitacao,
               r.solicitante_nome, r.funcionario_retirada, r.fornecedor,
               r.aprovador_nome, r.assinatura,
               ri.ordem, ri.descricao, ri.quantidade, ri.item_catalogo_id,
               ic.nome AS item_nome
        FROM requisicoes_compra r
        JOIN requisicoes_itens ri ON ri.requisicao_id = r.id
        LEFT JOIN itens_catalogo ic ON ic.id = ri.item_catalogo_id
        WHERE {where_sql}
        ORDER BY r.data_decisao DESC, r.id DESC, ri.ordem
    """, params).fetchall()

    resumo = db.execute(f"""
        SELECT ri.item_catalogo_id AS item_id,
               COALESCE(ic.nome, ri.descricao) AS item_nome,
               ic.unidade,
               COUNT(*) AS n_pedidos,
               MAX(r.data_decisao) AS ultima_autorizacao,
               COUNT(DISTINCT r.fornecedor) AS n_fornecedores,
               GROUP_CONCAT(DISTINCT r.fornecedor) AS fornecedores
        FROM requisicoes_compra r
        JOIN requisicoes_itens ri ON ri.requisicao_id = r.id
        LEFT JOIN itens_catalogo ic ON ic.id = ri.item_catalogo_id
        WHERE {where_sql}
        GROUP BY ri.item_catalogo_id, COALESCE(ic.nome, ri.descricao), ic.unidade
        ORDER BY n_pedidos DESC, item_nome
    """, params).fetchall()

    return {
        'filtros': {
            'data_ini': data_ini, 'data_fim': data_fim,
            'item_id': item_id, 'fornecedor': fornecedor,
        },
        'total_autorizacoes': len({r['req_id'] for r in linhas}),
        'total_itens': len(linhas),
        'linhas': [dict(r) for r in linhas],
        'resumo': [dict(r) for r in resumo],
    }


@app.route('/api/requisicoes/relatorio')
@api_login_required
def api_requisicoes_relatorio():
    data = _relatorio_query(get_db(), current_user(), request.args)
    data['is_master'] = is_master(current_user())
    return jsonify(data)


@app.route('/api/requisicoes/relatorio.csv')
@api_login_required
def api_requisicoes_relatorio_csv():
    data = _relatorio_query(get_db(), current_user(), request.args)
    output = io.StringIO()
    w = csv.writer(output, delimiter=';')
    w.writerow(['Data Autorização', 'Nº', 'Solicitante', 'Fornecedor',
                'Funcionário', 'Item', 'Quantidade', 'Autorizado por'])
    for r in data['linhas']:
        w.writerow([r['data_decisao'], r['numero'], r['solicitante_nome'], r['fornecedor'],
                    r['funcionario_retirada'], r['item_nome'] or r['descricao'],
                    r['quantidade'] or '', r['aprovador_nome']])
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=autorizacoes.csv'}
    )


# ── Backup do banco (somente master) ───────────────────────────────

@app.route('/admin/backup')
@master_required
def admin_backup():
    """Baixa um snapshot consistente do SQLite (online backup API)."""
    fd, tmp_path = tempfile.mkstemp(suffix='.db', prefix='fazenda_bkp_')
    os.close(fd)

    @after_this_request
    def _cleanup(response):
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        return response

    try:
        src = get_db()
        dst = sqlite3.connect(tmp_path)
        with dst:
            src.backup(dst)
        dst.close()
    except Exception as e:
        return f'Erro ao gerar backup: {e}', 500

    stamp = now_local().strftime('%Y%m%d_%H%M%S')
    return send_file(
        tmp_path,
        as_attachment=True,
        download_name=f'fazenda_backup_{stamp}.db',
        mimetype='application/octet-stream',
    )


# ── Admin: gerenciamento de usuários (master only) ─────────────────

@app.route('/admin/usuarios')
@master_required
def admin_usuarios_page():
    return render_template('admin_usuarios.html')


@app.route('/api/usuarios', methods=['GET'])
@api_master_required
def api_usuarios_list():
    db = get_db()
    rows = db.execute(
        "SELECT id, nome, email, papel, ativo, created_at FROM usuarios ORDER BY id"
    ).fetchall()
    return jsonify({
        'usuarios': [dict(r) for r in rows],
        'me': session.get('user_id'),
    })


@app.route('/api/usuarios', methods=['POST'])
@api_master_required
def api_usuarios_criar():
    db = get_db()
    data = request.get_json(silent=True) or {}
    nome  = (data.get('nome') or '').strip()
    email = (data.get('email') or '').strip().lower()
    senha = data.get('senha') or ''
    papel = (data.get('papel') or 'usuario').strip()

    if not nome:
        return jsonify({'erro': 'Informe o nome.'}), 400
    if not email:
        return jsonify({'erro': 'Informe o e-mail.'}), 400
    if len(senha) < 6:
        return jsonify({'erro': 'Senha deve ter ao menos 6 caracteres.'}), 400
    if papel not in ('usuario', 'master'):
        papel = 'usuario'

    if db.execute("SELECT 1 FROM usuarios WHERE lower(email)=?", (email,)).fetchone():
        return jsonify({'erro': 'Já existe um usuário com esse e-mail.'}), 400

    senha_hash = bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()
    cur = db.execute(
        "INSERT INTO usuarios (nome, email, senha_hash, papel, ativo) VALUES (?, ?, ?, ?, 1)",
        (nome, email, senha_hash, papel)
    )
    db.commit()
    return jsonify({'id': cur.lastrowid, 'ok': True}), 201


@app.route('/api/usuarios/<int:uid>/ativo', methods=['POST'])
@api_master_required
def api_usuarios_toggle_ativo(uid):
    db = get_db()
    if uid == session.get('user_id'):
        return jsonify({'erro': 'Você não pode desativar sua própria conta.'}), 400
    u = db.execute("SELECT ativo FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not u:
        return jsonify({'erro': 'Usuário não encontrado.'}), 404
    novo = 0 if u['ativo'] else 1
    db.execute("UPDATE usuarios SET ativo=? WHERE id=?", (novo, uid))
    db.commit()
    return jsonify({'ativo': novo})


@app.route('/api/usuarios/<int:uid>/senha', methods=['POST'])
@api_master_required
def api_usuarios_reset_senha(uid):
    db = get_db()
    data = request.get_json(silent=True) or {}
    senha = data.get('senha') or ''
    if len(senha) < 6:
        return jsonify({'erro': 'Senha deve ter ao menos 6 caracteres.'}), 400
    if not db.execute("SELECT 1 FROM usuarios WHERE id=?", (uid,)).fetchone():
        return jsonify({'erro': 'Usuário não encontrado.'}), 404
    senha_hash = bcrypt.hashpw(senha.encode(), bcrypt.gensalt()).decode()
    db.execute("UPDATE usuarios SET senha_hash=? WHERE id=?", (senha_hash, uid))
    db.commit()
    return jsonify({'ok': True})


@app.route('/api/usuarios/<int:uid>/papel', methods=['POST'])
@api_master_required
def api_usuarios_mudar_papel(uid):
    db = get_db()
    data = request.get_json(silent=True) or {}
    papel = (data.get('papel') or '').strip()
    if papel not in ('usuario', 'master'):
        return jsonify({'erro': 'Papel inválido.'}), 400
    if uid == session.get('user_id') and papel != 'master':
        return jsonify({'erro': 'Você não pode remover seu próprio papel de master.'}), 400
    if not db.execute("SELECT 1 FROM usuarios WHERE id=?", (uid,)).fetchone():
        return jsonify({'erro': 'Usuário não encontrado.'}), 404
    db.execute("UPDATE usuarios SET papel=? WHERE id=?", (papel, uid))
    db.commit()
    return jsonify({'ok': True, 'papel': papel})


# ── API: Embriões ──────────────────────────────────────────────────

@app.route('/api/embrioes', methods=['GET'])
@api_login_required
def api_embrioes_lista():
    db = get_db()
    q = (request.args.get('q') or '').strip()
    tipo = request.args.get('tipo') or ''
    touro = (request.args.get('touro') or '').strip()
    zerados = request.args.get('zerados') == '1'

    sql = "SELECT * FROM embriao_lote WHERE 1=1"
    params = []
    if q:
        sql += " AND (doadora LIKE ? OR touro LIKE ?)"
        params += [f"%{q}%", f"%{q}%"]
    if tipo in ('Sex F', 'Conv.'):
        sql += " AND tipo_semen = ?"
        params.append(tipo)
    if touro:
        sql += " AND touro = ?"
        params.append(touro)
    if not zerados:
        sql += " AND qtd_atual > 0"
    sql += " ORDER BY dt_vitrificacao DESC, id DESC"

    rows = [dict(r) for r in db.execute(sql, params).fetchall()]
    return jsonify(rows)


@app.route('/api/embrioes/kpis', methods=['GET'])
@api_login_required
def api_embrioes_kpis():
    db = get_db()
    row = db.execute("""
        SELECT
            COALESCE(SUM(qtd_atual), 0)                                    AS total,
            COALESCE(SUM(CASE WHEN tipo_semen='Sex F' THEN qtd_atual END), 0) AS sex_f,
            COALESCE(SUM(CASE WHEN tipo_semen='Conv.' THEN qtd_atual END), 0) AS conv,
            COUNT(DISTINCT CASE WHEN qtd_atual > 0 THEN doadora END)        AS doadoras,
            COUNT(DISTINCT CASE WHEN qtd_atual > 0 THEN touro END)          AS touros
        FROM embriao_lote
    """).fetchone()
    receita = db.execute(
        "SELECT COALESCE(SUM(valor_total), 0) AS r "
        "FROM embriao_movimento WHERE tipo='venda'"
    ).fetchone()
    return jsonify({
        "total": row["total"],
        "sex_f": row["sex_f"],
        "conv": row["conv"],
        "doadoras": row["doadoras"],
        "touros": row["touros"],
        "receita_total": receita["r"],
    })


@app.route('/api/embrioes', methods=['POST'])
@api_login_required
def api_embrioes_criar():
    data = request.json or {}
    required = ['doadora', 'touro', 'tipo_semen', 'qtd']
    for k in required:
        if not data.get(k):
            return jsonify({'erro': f'Campo obrigatório: {k}'}), 400
    try:
        qtd = int(data['qtd'])
    except (TypeError, ValueError):
        return jsonify({'erro': 'qtd deve ser inteiro'}), 400
    if qtd <= 0:
        return jsonify({'erro': 'qtd deve ser positivo'}), 400
    tipo = data['tipo_semen']
    if tipo not in ('Sex F', 'Conv.'):
        return jsonify({'erro': 'tipo_semen inválido'}), 400

    doadora = data['doadora'].strip()
    db = get_db()
    try:
        cur = db.execute("""
            INSERT INTO embriao_lote
              (dt_opu, dt_vitrificacao, doadora, doadora_matriz_id,
               touro, tipo_semen, qtd_inicial, qtd_atual, obs, arquivo_origem)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            (data.get('dt_opu') or '').strip(),
            (data.get('dt_vitrificacao') or '').strip(),
            doadora,
            lookup_matriz(doadora),
            data['touro'].strip().upper(),
            tipo,
            qtd, qtd,
            (data.get('obs') or '').strip() or None,
            'manual',
        ))
        db.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Lote já existe (mesma chave única)'}), 409


@app.route('/api/embrioes/<int:lote_id>', methods=['GET'])
@api_login_required
def api_embrioes_detalhe(lote_id):
    db = get_db()
    lote = db.execute(
        "SELECT * FROM embriao_lote WHERE id=?", (lote_id,)
    ).fetchone()
    if not lote:
        return jsonify({'erro': 'Lote não encontrado'}), 404


    movs = [dict(r) for r in db.execute(
        "SELECT m.*, u.nome AS user_nome FROM embriao_movimento m "
        "LEFT JOIN usuarios u ON u.id=m.created_by "
        "WHERE m.lote_id=? ORDER BY m.data, m.id", (lote_id,)
    ).fetchall()]

    agg = db.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN tipo='te_interna' THEN qtd END), 0) AS usado_te,
            COALESCE(SUM(CASE WHEN tipo='venda'       THEN qtd END), 0) AS vendido,
            COALESCE(SUM(CASE WHEN tipo='venda'       THEN valor_total END), 0) AS receita
        FROM embriao_movimento WHERE lote_id=?
    """, (lote_id,)).fetchone()

    return jsonify({
        'lote': dict(lote),
        'movimentos': movs,
        'kpis': {
            'restante': lote['qtd_atual'],
            'usado_te': agg['usado_te'],
            'vendido': agg['vendido'],
            'receita': agg['receita'],
        },
    })


@app.route('/api/embrioes/<int:lote_id>', methods=['PUT'])
@api_master_required
def api_embrioes_editar(lote_id):
    data = request.json or {}
    # Whitelist of editable fields (qtd_atual deliberately excluded)
    editable = ['dt_opu', 'dt_vitrificacao', 'doadora', 'touro',
                'tipo_semen', 'obs']
    sets, params = [], []
    for k in editable:
        if k in data:
            v = (data[k] or '').strip() if isinstance(data[k], str) else data[k]
            if k == 'touro' and v:
                v = v.upper()
            if k == 'tipo_semen' and v not in ('Sex F', 'Conv.'):
                return jsonify({'erro': 'tipo_semen inválido'}), 400
            sets.append(f"{k}=?")
            params.append(v if v != '' else None)
    if not sets:
        return jsonify({'ok': True})
    # Re-link doadora if it changed
    if 'doadora' in data:
        sets.append("doadora_matriz_id=?")
        params.append(lookup_matriz(data['doadora']))
    params.append(lote_id)
    db = get_db()
    try:
        db.execute(f"UPDATE embriao_lote SET {','.join(sets)} WHERE id=?", params)
        db.commit()
    except sqlite3.IntegrityError as e:
        return jsonify({'erro': f'Conflito de unicidade: {e}'}), 409
    return jsonify({'ok': True})


@app.route('/api/embrioes/<int:lote_id>', methods=['DELETE'])
@api_master_required
def api_embrioes_excluir(lote_id):
    db = get_db()
    cur = db.execute("DELETE FROM embriao_lote WHERE id=?", (lote_id,))
    db.commit()
    if cur.rowcount == 0:
        return jsonify({'erro': 'Lote não encontrado'}), 404
    return jsonify({'ok': True})


@app.route('/api/embrioes/<int:lote_id>/movimento', methods=['POST'])
@api_login_required
def api_embrioes_criar_movimento(lote_id):
    data = request.json or {}
    tipo = data.get('tipo')
    if tipo not in ('te_interna', 'venda', 'perda', 'ajuste_lab'):
        return jsonify({'erro': 'Tipo inválido'}), 400
    if tipo == 'ajuste_lab' and not is_master(current_user()):
        return jsonify({'erro': 'Apenas master pode criar ajuste de reconciliação'}), 403

    try:
        qtd = int(data.get('qtd', 0))
    except (TypeError, ValueError):
        return jsonify({'erro': 'qtd deve ser inteiro'}), 400
    if qtd <= 0:
        return jsonify({'erro': 'qtd deve ser positivo'}), 400

    data_mov = (data.get('data') or '').strip()
    if not data_mov:
        return jsonify({'erro': 'Campo data é obrigatório'}), 400

    valor_unit = data.get('valor_unit')
    if valor_unit is not None:
        try:
            valor_unit = float(valor_unit)
        except (TypeError, ValueError):
            return jsonify({'erro': 'valor_unit inválido'}), 400

    db = get_db()
    db.execute('BEGIN')
    try:
        lote = db.execute(
            "SELECT qtd_atual FROM embriao_lote WHERE id=?", (lote_id,)
        ).fetchone()
        if not lote:
            db.rollback()
            return jsonify({'erro': 'Lote não encontrado'}), 404
        if qtd > lote['qtd_atual']:
            db.rollback()
            return jsonify({
                'erro': f'Saldo insuficiente (disponível: {lote["qtd_atual"]})'
            }), 400

        valor_total = (valor_unit * qtd) if (tipo == 'venda' and valor_unit) else None

        db.execute("""
            INSERT INTO embriao_movimento
              (lote_id, tipo, qtd, data, receptora, comprador,
               valor_unit, valor_total, obs, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            lote_id, tipo, qtd, data_mov,
            (data.get('receptora') or '').strip() or None,
            (data.get('comprador') or '').strip() or None,
            valor_unit, valor_total,
            (data.get('obs') or '').strip() or None,
            session.get('user_id'),
        ))
        db.execute(
            "UPDATE embriao_lote SET qtd_atual = qtd_atual - ? WHERE id=?",
            (qtd, lote_id)
        )
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500


@app.route('/api/embrioes/movimento/<int:mov_id>', methods=['DELETE'])
@api_login_required
def api_embrioes_excluir_movimento(mov_id):
    db = get_db()
    mov = db.execute(
        "SELECT lote_id, qtd, tipo FROM embriao_movimento WHERE id=?",
        (mov_id,)
    ).fetchone()
    if not mov:
        return jsonify({'erro': 'Movimento não encontrado'}), 404
    if mov['tipo'] == 'ajuste_lab' and not is_master(current_user()):
        return jsonify({
            'erro': 'Apenas master pode excluir ajuste de reconciliação'
        }), 403

    db.execute('BEGIN')
    try:
        db.execute(
            "UPDATE embriao_lote SET qtd_atual = qtd_atual + ? WHERE id=?",
            (mov['qtd'], mov['lote_id'])
        )
        db.execute("DELETE FROM embriao_movimento WHERE id=?", (mov_id,))
        db.commit()
        return jsonify({'ok': True})
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500


@app.route('/api/embrioes/importar/preview', methods=['POST'])
@api_login_required
def api_embrioes_importar_preview():
    """Parse uploaded PDF and return the extracted rows without persisting."""
    if 'arquivo' not in request.files:
        return jsonify({'erro': 'arquivo PDF não enviado'}), 400
    f = request.files['arquivo']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'erro': 'apenas arquivos PDF são aceitos'}), 400
    # Persist to a temp path so pdfplumber can open by name
    tmp = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
    try:
        f.save(tmp.name)
        tmp.close()
        result = parse_fivet_pdf(tmp.name)
        result['arquivo'] = f.filename
        return jsonify(result)
    except ValueError as e:
        return jsonify({'erro': str(e)}), 400
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


@app.route('/api/embrioes/importar/confirmar', methods=['POST'])
@api_login_required
def api_embrioes_importar_confirmar():
    """Persist the (possibly user-edited) rows. UNIQUE constraint dedupes."""
    payload = request.json or {}
    linhas = payload.get('linhas') or []
    if not linhas:
        return jsonify({'erro': 'Nenhuma linha para importar'}), 400
    arquivo = (payload.get('arquivo') or '').strip()
    data_planilha = (payload.get('data_planilha') or '').strip() or None

    db = get_db()
    novos, ignorados = 0, 0
    db.execute('BEGIN')
    try:
        for row in linhas:
            try:
                qtd = int(row.get('qtd', 0))
            except (TypeError, ValueError):
                continue
            if qtd <= 0:
                continue
            doadora = (row.get('doadora') or '').strip()
            touro = (row.get('touro') or '').strip().upper()
            tipo = row.get('tipo_semen') or 'Conv.'
            if tipo not in ('Sex F', 'Conv.'):
                tipo = 'Conv.'
            if not doadora or not touro:
                continue
            try:
                db.execute("""
                    INSERT INTO embriao_lote
                      (dt_opu, dt_vitrificacao, doadora, doadora_matriz_id,
                       touro, tipo_semen, qtd_inicial, qtd_atual, obs, arquivo_origem)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (
                    (row.get('dt_opu') or '').strip(),
                    (row.get('dt_vitrificacao') or '').strip(),
                    doadora,
                    lookup_matriz(doadora),
                    touro, tipo, qtd, qtd,
                    (row.get('obs') or '').strip() or None,
                    arquivo or None,
                ))
                novos += 1
            except sqlite3.IntegrityError:
                ignorados += 1
        total = sum(int(r.get('qtd', 0)) for r in linhas if str(r.get('qtd', '')).isdigit())
        db.execute("""
            INSERT INTO embriao_import
              (arquivo, data_planilha, n_lotes_novos, n_lotes_ignorados,
               n_embrioes_total, imported_by)
            VALUES (?,?,?,?,?,?)
        """, (arquivo or None, data_planilha, novos, ignorados, total,
              session.get('user_id')))
        db.commit()
        return jsonify({'ok': True, 'novos': novos, 'ignorados': ignorados,
                        'total': total})
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500


@app.route('/api/embrioes/reconciliar/preview', methods=['POST'])
@api_master_required
def api_embrioes_reconciliar_preview():
    """Compare DB lotes vs PDF rows. Returns 3 lists: novos, sumidos, divergentes."""
    payload = request.json or {}
    linhas = payload.get('linhas') or []
    db = get_db()
    # Build lookup of planilha rows by unique key
    chave = lambda r: (
        (r.get('dt_opu') or '').strip(),
        (r.get('dt_vitrificacao') or '').strip(),
        (r.get('doadora') or '').strip(),
        (r.get('touro') or '').strip().upper(),
        r.get('tipo_semen') or 'Conv.',
    )
    pdf_map = {chave(r): int(r.get('qtd', 0)) for r in linhas}

    db_rows = db.execute(
        "SELECT id, dt_opu, dt_vitrificacao, doadora, touro, tipo_semen, qtd_atual "
        "FROM embriao_lote"
    ).fetchall()

    novos, sumidos, divergentes = [], [], []
    db_keys = set()
    for r in db_rows:
        k = ((r['dt_opu'] or '').strip(),
             (r['dt_vitrificacao'] or '').strip(),
             (r['doadora'] or '').strip(),
             (r['touro'] or '').strip().upper(),
             r['tipo_semen'])
        db_keys.add(k)
        if k in pdf_map:
            if pdf_map[k] != r['qtd_atual']:
                divergentes.append({
                    'lote_id': r['id'],
                    'doadora': r['doadora'],
                    'touro': r['touro'],
                    'tipo_semen': r['tipo_semen'],
                    'dt_vitrificacao': r['dt_vitrificacao'],
                    'sistema': r['qtd_atual'],
                    'planilha': pdf_map[k],
                    'diferenca': pdf_map[k] - r['qtd_atual'],
                })
        elif r['qtd_atual'] > 0:
            sumidos.append({
                'lote_id': r['id'],
                'doadora': r['doadora'],
                'touro': r['touro'],
                'tipo_semen': r['tipo_semen'],
                'dt_vitrificacao': r['dt_vitrificacao'],
                'sistema': r['qtd_atual'],
            })

    for k, qtd_pdf in pdf_map.items():
        if k not in db_keys and qtd_pdf > 0:
            novos.append({
                'dt_opu': k[0], 'dt_vitrificacao': k[1],
                'doadora': k[2], 'touro': k[3],
                'tipo_semen': k[4], 'qtd': qtd_pdf,
            })

    return jsonify({'novos': novos, 'sumidos': sumidos, 'divergentes': divergentes})


@app.route('/api/embrioes/reconciliar/aplicar', methods=['POST'])
@api_master_required
def api_embrioes_reconciliar_aplicar():
    """Apply selected adjustments. Creates ajuste_lab movements."""
    payload = request.json or {}
    ajustes = payload.get('ajustes') or []
    data_planilha = (payload.get('data_planilha') or '').strip() or '?'

    db = get_db()
    db.execute('BEGIN')
    aplicados = 0
    try:
        # Note: `novos` entries from preview (lotes in PDF but not in DB) cannot
        # be applied here — they have no lote_id. Use /api/embrioes/importar/confirmar
        # to insert new lotes from the PDF.
        for adj in ajustes:
            lote_id = adj.get('lote_id')
            if not lote_id:
                continue
            try:
                diff = int(adj.get('diferenca', 0))
            except (TypeError, ValueError):
                continue
            if diff == 0:
                continue
            if diff > 0:
                # Sistema tinha menos que planilha — devolve embriões
                db.execute(
                    "UPDATE embriao_lote SET qtd_atual = qtd_atual + ? WHERE id=?",
                    (diff, lote_id)
                )
                qtd_mov = diff
                obs = f"Reconciliação com planilha {data_planilha}: sistema +{diff}."
            else:
                qtd_mov = -diff  # positivo
                lote = db.execute(
                    "SELECT qtd_atual FROM embriao_lote WHERE id=?", (lote_id,)
                ).fetchone()
                if not lote or qtd_mov > lote['qtd_atual']:
                    continue  # skip if would go negative
                db.execute(
                    "UPDATE embriao_lote SET qtd_atual = qtd_atual - ? WHERE id=?",
                    (qtd_mov, lote_id)
                )
                obs = f"Reconciliação com planilha {data_planilha}: sistema -{qtd_mov}."
            db.execute("""
                INSERT INTO embriao_movimento
                  (lote_id, tipo, qtd, data, obs, created_by)
                VALUES (?, 'ajuste_lab', ?, ?, ?, ?)
            """, (lote_id, qtd_mov, data_planilha, obs, session.get('user_id')))
            aplicados += 1
        db.commit()
        return jsonify({'ok': True, 'aplicados': aplicados})
    except Exception as e:
        db.rollback()
        return jsonify({'erro': str(e)}), 500


@app.route('/api/embrioes/doadora-info/<path:doadora>', methods=['GET'])
@api_login_required
def api_embrioes_doadora_info(doadora):
    matched = lookup_matriz(doadora)
    if not matched:
        return jsonify({'erro': 'doadora não cadastrada como matriz'}), 404
    db = get_db()
    m = db.execute("""
        SELECT m.animal_id, m.categoria, m.ceip, m.precoce,
               a.iciagen, a.deca_icia_g, a.idesm, a.rmat
        FROM matrizes m
        LEFT JOIN avaliacoes a ON a.animal_id = m.animal_id
        WHERE m.animal_id = ?
        ORDER BY a.rodada_id DESC LIMIT 1
    """, (matched,)).fetchone()
    return jsonify(dict(m))


@app.route('/api/embrioes/relink', methods=['POST'])
@api_master_required
def api_embrioes_relink():
    """Re-runs lookup_matriz for every lote and updates doadora_matriz_id."""
    db = get_db()
    rows = db.execute("SELECT id, doadora FROM embriao_lote").fetchall()
    atualizados = 0
    for r in rows:
        novo = lookup_matriz(r['doadora'])
        db.execute(
            "UPDATE embriao_lote SET doadora_matriz_id=? WHERE id=?",
            (novo, r['id'])
        )
        atualizados += 1
    db.commit()
    return jsonify({'ok': True, 'atualizados': atualizados})


# ── PecIA: Chat conversacional com Claude (tool use, read-only) ────

PECIA_SYSTEM_PROMPT = """Você é o PecIA, assistente de pecuária da Fazenda Porto do Engenho — cabanha de Nelore Provado em Cariacica-ES (CIA de Melhoramento). Responde perguntas sobre o rebanho consultando o banco de dados via ferramentas.

CONTEXTO DO REBANHO:
- Sistema de melhoramento genético baseado em rodadas de avaliação importadas periodicamente. Os dados refletem a rodada mais recente.
- Índices principais (quanto MAIOR melhor): ICIAGen (mérito genético geral), IDESM (índice desmama), RMAT (mérito reprodutivo), IFRIG (índice frigorífico).
- IPP (Idade ao Primeiro Parto) e IEP (Intervalo Entre Partos): em meses, quanto MENOR melhor.
- DECA = decil dentro do rebanho. DECA 1 = top 10% (excelente), DECA 10 = pior 10% (candidato a descarte).
- Categorias: M=Matriz (já pariu), N=Novilha. CEIP=Certificado de Eficiência Produtiva (PMGRN-Embrapa).
- Estoque FIV: lotes de embriões por combinação doadora×touro, com qtd_atual disponível.
- Estoque de touros para venda: machos no inventário comercial, agrupados por safra.

REGRAS:
1. SEMPRE use as ferramentas para responder qualquer pergunta factual sobre animais, índices ou estoque. NUNCA invente brincos, números ou dados.
2. Se uma ferramenta retornar vazio/não encontrado, diga ao usuário que não achou — não chute.
3. Brincos podem ter formatos variados (ex: "0001", "54 KO11", "BRA12345"). Passe exatamente como o usuário digitou.
4. Seja objetivo. Listas de animais: use tabela com 3-5 colunas essenciais.
5. Quando o usuário pedir "as melhores" sem dizer qual índice, assuma ICIAGen.
6. Você é READ-ONLY. Se pedirem alteração de dados, oriente onde fazer pelo próprio sistema.
7. Use português brasileiro objetivo, sem jargão excessivo."""


PECIA_TOOLS = [
    {
        "name": "consultar_animal",
        "description": "Ficha completa de um animal pelo brinco/ID: sexo, categoria (M=matriz/N=novilha), pai, mãe, avós, data de nascimento, índices (ICIAGen, IDESM, RMAT, IFRIG), DECAs, IPP, IEP, CEIP, status de genotipagem, e onde está cadastrado (matriz ativa, estoque de venda, ou produto/cria). Use sempre que o usuário perguntar sobre um animal específico.",
        "input_schema": {
            "type": "object",
            "properties": {
                "brinco": {"type": "string", "description": "Brinco/ID do animal exatamente como o usuário digitou"}
            },
            "required": ["brinco"]
        }
    },
    {
        "name": "genealogia",
        "description": "Genealogia de um animal: pai (touro), mãe, avós paterno e materno, e lista de filhos (cruzando matrizes onde touro_pai/mae_id = brinco, e produtos onde touro/mae_id = brinco). Use para 'quem é filho de X?', 'pais de Y?', 'quantos filhos X tem?'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "brinco": {"type": "string", "description": "Brinco do animal"}
            },
            "required": ["brinco"]
        }
    },
    {
        "name": "top_animais_por_indice",
        "description": "Ranking dos N melhores (ou piores) animais por um índice genético. Use para 'top 10 ICIAGen', 'minhas melhores vacas', 'piores IEP', 'novilhas com melhor IDESM'. A função já considera qual direção (maior ou menor) é 'melhor' para cada índice.",
        "input_schema": {
            "type": "object",
            "properties": {
                "indice": {"type": "string", "enum": ["iciagen", "idesm", "rmat", "ifrig", "iep", "ipp"], "description": "Índice a ranquear"},
                "categoria": {"type": "string", "enum": ["M", "N", "TODAS"], "description": "M=Matriz, N=Novilha, TODAS=ambas. Default: TODAS."},
                "limit": {"type": "integer", "description": "Quantos retornar. Default 10, máximo 50."},
                "ordem": {"type": "string", "enum": ["melhores", "piores"], "description": "Default 'melhores'."}
            },
            "required": ["indice"]
        }
    },
    {
        "name": "estoque_embrioes",
        "description": "Lista lotes de embriões FIV no estoque. Filtre por doadora (vaca de origem) e/ou touro. Retorna doadora, touro, qtd_atual, qtd_inicial, datas de OPU/vitrificação, laboratório. Use para 'quanto embrião eu tenho?', 'embrião da vaca X', 'embrião do touro Y'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "doadora": {"type": "string", "description": "Filtrar por doadora (busca parcial). Opcional."},
                "touro": {"type": "string", "description": "Filtrar por touro (busca parcial). Opcional."},
                "apenas_disponiveis": {"type": "boolean", "description": "Default true: só retorna lotes com qtd_atual>0."}
            }
        }
    },
    {
        "name": "estoque_touros_venda",
        "description": "Lista touros do estoque comercial para venda. Mostra brinco, índices (ICIAGen, IDESM, RMAT, IFRIG), peso, pai, grupo/safra e status (vendido/disponível). Use para 'touros pra vender', 'estoque safra X', 'melhores touros disponíveis'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "grupo": {"type": "string", "description": "Nome do grupo/safra. Opcional."},
                "apenas_disponiveis": {"type": "boolean", "description": "Default true: só não vendidos."}
            }
        }
    },
    {
        "name": "resumo_rebanho",
        "description": "Visão geral do rebanho na rodada mais recente: total de matrizes ativas, médias de ICIAGen/IDESM/RMAT/IPP/IEP, distribuição por categoria, totais de CEIP/genotipadas/precoces, estoque de embriões e touros à venda. Use para 'panorama geral', 'como tá meu rebanho?', 'quantas matrizes eu tenho?'.",
        "input_schema": {"type": "object", "properties": {}}
    },
    {
        "name": "gerar_relatorio_pdf",
        "description": "Gera um relatório em PDF e retorna a URL para download (não envia arquivo no chat — o link aparece como anexo). Use sempre que o usuário pedir 'PDF', 'relatório', 'exportar', 'baixar', 'imprimir'. Tipos disponíveis: 'estoque_touros' (touros à venda, paisagem) ou 'estoque_embrioes' (lotes FIV, retrato). Filtros são opcionais.",
        "input_schema": {
            "type": "object",
            "properties": {
                "tipo": {
                    "type": "string",
                    "enum": ["estoque_touros", "estoque_embrioes"],
                    "description": "Tipo do relatório."
                },
                "apenas_disponiveis": {
                    "type": "boolean",
                    "description": "Default true. estoque_touros: só não vendidos. estoque_embrioes: só lotes com saldo > 0."
                },
                "filtro_grupo": {
                    "type": "string",
                    "description": "Só para estoque_touros: filtra por grupo/safra (busca parcial)."
                },
                "filtro_doadora": {
                    "type": "string",
                    "description": "Só para estoque_embrioes: filtra por doadora (busca parcial)."
                },
                "filtro_touro": {
                    "type": "string",
                    "description": "Só para estoque_embrioes: filtra por touro (busca parcial)."
                }
            },
            "required": ["tipo"]
        }
    }
]


def _pecia_tool_consultar_animal(db, brinco):
    brinco = (brinco or '').strip()
    if not brinco:
        return {"erro": "Brinco vazio"}
    rodada = get_ultima_rodada(db)
    if not rodada:
        return {"erro": "Nenhuma rodada importada"}
    rid = rodada['id']

    m = db.execute("""
        SELECT m.animal_id, m.data_nasc, m.touro_pai, m.mae_id, m.avo_paterno, m.avo_materno,
               m.categoria, m.genotipada, m.precoce, m.ceip, m.ipp, m.iep, m.pv, m.desc_ap,
               a.iciagen, a.deca_icia_g, a.idesm, a.deca_idesm_g, a.rmat, a.deca_rmat, a.ifrig
        FROM matrizes m
        LEFT JOIN avaliacoes a ON a.animal_id=m.animal_id AND a.rodada_id=m.rodada_id
        WHERE m.animal_id = ? AND m.rodada_id = ? AND m.ativo = 1
    """, (brinco, rid)).fetchone()
    if m:
        return {
            "encontrado_em": "matriz_ativa",
            "brinco": m['animal_id'], "sexo": "F", "categoria": m['categoria'],
            "data_nascimento": m['data_nasc'],
            "pai": m['touro_pai'], "mae": m['mae_id'],
            "avo_paterno": m['avo_paterno'], "avo_materno": m['avo_materno'],
            "iciagen": m['iciagen'], "deca_iciagen": m['deca_icia_g'],
            "idesm": m['idesm'], "deca_idesm": m['deca_idesm_g'],
            "rmat": m['rmat'], "deca_rmat": m['deca_rmat'], "ifrig": m['ifrig'],
            "ipp_meses": m['ipp'], "iep_meses": m['iep'], "peso_vivo": m['pv'],
            "genotipada": bool(m['genotipada']), "precoce": bool(m['precoce']),
            "ceip": bool(m['ceip']), "obs_aprumos": m['desc_ap']
        }

    t = db.execute("""
        SELECT et.brinco, et.data_nasc, et.pai, et.avo_paterno, et.avo_materno,
               et.idesm, et.iciagen, et.rmat, et.ifrig, et.peso, et.vendido,
               eg.nome AS grupo
        FROM estoque_touros et
        LEFT JOIN estoque_grupos eg ON eg.id = et.grupo_id
        WHERE et.brinco = ?
    """, (brinco,)).fetchone()
    if t:
        return {
            "encontrado_em": "estoque_venda",
            "brinco": t['brinco'], "sexo": "M",
            "data_nascimento": t['data_nasc'], "pai": t['pai'],
            "avo_paterno": t['avo_paterno'], "avo_materno": t['avo_materno'],
            "iciagen": t['iciagen'], "idesm": t['idesm'],
            "rmat": t['rmat'], "ifrig": t['ifrig'],
            "peso_kg": t['peso'], "grupo_safra": t['grupo'],
            "vendido": bool(t['vendido'])
        }

    p = db.execute("""
        SELECT produto_id, mae_id, touro, sexo, data_nasc, pn, peso_desm,
               idesm, iciagen, rmat
        FROM produtos WHERE produto_id = ? AND rodada_id = ?
    """, (brinco, rid)).fetchone()
    if p:
        return {
            "encontrado_em": "produto_safra",
            "brinco": p['produto_id'], "sexo": p['sexo'],
            "data_nascimento": p['data_nasc'], "pai": p['touro'], "mae": p['mae_id'],
            "iciagen": p['iciagen'], "idesm": p['idesm'], "rmat": p['rmat'],
            "peso_nascimento": p['pn'], "peso_desmama": p['peso_desm']
        }

    return {"encontrado": False, "brinco_buscado": brinco,
            "mensagem": "Não encontrado em matrizes ativas, estoque de venda nem produtos da rodada atual."}


def _pecia_tool_genealogia(db, brinco):
    brinco = (brinco or '').strip()
    if not brinco:
        return {"erro": "Brinco vazio"}
    rodada = get_ultima_rodada(db)
    if not rodada:
        return {"erro": "Nenhuma rodada importada"}
    rid = rodada['id']

    out = {"brinco": brinco, "pais": None}

    m = db.execute("""
        SELECT touro_pai, mae_id, avo_paterno, avo_materno
        FROM matrizes WHERE animal_id=? AND rodada_id=? AND ativo=1
    """, (brinco, rid)).fetchone()
    if m:
        out["pais"] = {"pai": m['touro_pai'], "mae": m['mae_id'],
                       "avo_paterno": m['avo_paterno'], "avo_materno": m['avo_materno']}

    filhos_pai_mat = db.execute("""
        SELECT animal_id, categoria FROM matrizes
        WHERE touro_pai=? AND rodada_id=? AND ativo=1
        ORDER BY animal_id LIMIT 100
    """, (brinco, rid)).fetchall()
    filhos_pai_prod = db.execute("""
        SELECT produto_id, sexo FROM produtos
        WHERE touro=? AND rodada_id=?
        ORDER BY produto_id LIMIT 100
    """, (brinco, rid)).fetchall()
    out["filhos_como_pai"] = (
        [{"brinco": r['animal_id'], "categoria": r['categoria'], "tabela": "matrizes"} for r in filhos_pai_mat]
        + [{"brinco": r['produto_id'], "sexo": r['sexo'], "tabela": "produtos"} for r in filhos_pai_prod]
    )

    filhos_mae_mat = db.execute("""
        SELECT animal_id, categoria FROM matrizes
        WHERE mae_id=? AND rodada_id=? AND ativo=1
        ORDER BY animal_id LIMIT 100
    """, (brinco, rid)).fetchall()
    filhos_mae_prod = db.execute("""
        SELECT produto_id, sexo FROM produtos
        WHERE mae_id=? AND rodada_id=?
        ORDER BY produto_id LIMIT 100
    """, (brinco, rid)).fetchall()
    out["filhos_como_mae"] = (
        [{"brinco": r['animal_id'], "categoria": r['categoria'], "tabela": "matrizes"} for r in filhos_mae_mat]
        + [{"brinco": r['produto_id'], "sexo": r['sexo'], "tabela": "produtos"} for r in filhos_mae_prod]
    )

    out["total_filhos_como_pai"] = len(out["filhos_como_pai"])
    out["total_filhos_como_mae"] = len(out["filhos_como_mae"])

    if not out["pais"] and not out["filhos_como_pai"] and not out["filhos_como_mae"]:
        return {"encontrado": False, "brinco_buscado": brinco}
    return out


def _pecia_tool_top_indice(db, indice, categoria='TODAS', limit=10, ordem='melhores'):
    rodada = get_ultima_rodada(db)
    if not rodada:
        return {"erro": "Nenhuma rodada importada"}
    rid = rodada['id']

    indices_avaliacoes = {'iciagen', 'idesm', 'rmat', 'ifrig'}
    indices_matrizes = {'iep', 'ipp'}
    if indice not in indices_avaliacoes and indice not in indices_matrizes:
        return {"erro": f"Índice '{indice}' não suportado"}

    melhor_eh_menor = indice in {'iep', 'ipp'}
    direcao = 'ASC' if ((ordem == 'melhores') == melhor_eh_menor) else 'DESC'

    try:
        limit = max(1, min(int(limit or 10), 50))
    except (TypeError, ValueError):
        limit = 10

    cat_filter = ''
    params = [rid]
    if categoria and categoria != 'TODAS':
        cat_filter = " AND m.categoria = ?"
        params.append(categoria)

    if indice in indices_avaliacoes:
        sql = f"""
            SELECT m.animal_id, m.categoria, a.{indice} AS valor,
                   a.iciagen, m.touro_pai, m.mae_id
            FROM matrizes m
            JOIN avaliacoes a ON a.animal_id=m.animal_id AND a.rodada_id=m.rodada_id
            WHERE m.rodada_id=? AND m.ativo=1 AND a.{indice} IS NOT NULL{cat_filter}
            ORDER BY a.{indice} {direcao} LIMIT ?
        """
    else:
        sql = f"""
            SELECT m.animal_id, m.categoria, m.{indice} AS valor,
                   a.iciagen, m.touro_pai, m.mae_id
            FROM matrizes m
            LEFT JOIN avaliacoes a ON a.animal_id=m.animal_id AND a.rodada_id=m.rodada_id
            WHERE m.rodada_id=? AND m.ativo=1 AND m.{indice} IS NOT NULL{cat_filter}
            ORDER BY m.{indice} {direcao} LIMIT ?
        """
    params.append(limit)
    rows = db.execute(sql, params).fetchall()
    return {
        "indice": indice, "ordem": ordem, "categoria_filtro": categoria,
        "total_retornado": len(rows),
        "ranking": [dict(r) for r in rows]
    }


def _pecia_tool_estoque_embrioes(db, doadora=None, touro=None, apenas_disponiveis=True):
    where = []
    params = []
    if doadora:
        where.append("doadora LIKE ?")
        params.append(f"%{doadora}%")
    if touro:
        where.append("touro LIKE ?")
        params.append(f"%{touro}%")
    if apenas_disponiveis:
        where.append("qtd_atual > 0")
    sql = ("SELECT id, dt_opu, dt_vitrificacao, doadora, touro, tipo_semen, "
           "qtd_inicial, qtd_atual, lab FROM embriao_lote")
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY qtd_atual DESC, dt_vitrificacao DESC LIMIT 200"
    rows = db.execute(sql, params).fetchall()
    total_disp = sum((r['qtd_atual'] or 0) for r in rows)
    return {
        "total_lotes": len(rows),
        "total_embrioes_disponiveis": total_disp,
        "filtros": {"doadora": doadora, "touro": touro, "apenas_disponiveis": apenas_disponiveis},
        "lotes": [dict(r) for r in rows]
    }


def _pecia_tool_estoque_touros_venda(db, grupo=None, apenas_disponiveis=True):
    where = []
    params = []
    if grupo:
        where.append("eg.nome LIKE ?")
        params.append(f"%{grupo}%")
    if apenas_disponiveis:
        where.append("et.vendido = 0")
    sql = """
        SELECT et.brinco, eg.nome AS grupo, et.data_nasc, et.pai,
               et.iciagen, et.idesm, et.rmat, et.ifrig, et.peso, et.vendido
        FROM estoque_touros et
        LEFT JOIN estoque_grupos eg ON eg.id = et.grupo_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY et.iciagen DESC NULLS LAST LIMIT 200"
    rows = db.execute(sql, params).fetchall()
    return {
        "total": len(rows),
        "filtros": {"grupo": grupo, "apenas_disponiveis": apenas_disponiveis},
        "touros": [dict(r) for r in rows]
    }


def _pecia_tool_resumo_rebanho(db):
    rodada = get_ultima_rodada(db)
    if not rodada:
        return {"erro": "Nenhuma rodada importada"}
    rid = rodada['id']
    kpis = db.execute("""
        SELECT COUNT(*) AS total_matrizes,
               AVG(a.iciagen) AS iciagen_avg,
               AVG(a.idesm) AS idesm_avg,
               AVG(a.rmat) AS rmat_avg,
               AVG(m.ipp) AS ipp_avg,
               AVG(m.iep) AS iep_avg,
               SUM(m.ceip) AS ceip_total,
               SUM(m.genotipada) AS genotipadas,
               SUM(m.precoce) AS precoces
        FROM matrizes m
        LEFT JOIN avaliacoes a ON a.animal_id=m.animal_id AND a.rodada_id=m.rodada_id
        WHERE m.rodada_id=? AND m.ativo=1
    """, (rid,)).fetchone()
    cats = db.execute("""
        SELECT categoria, COUNT(*) AS n FROM matrizes
        WHERE rodada_id=? AND ativo=1 GROUP BY categoria
    """, (rid,)).fetchall()
    emb = db.execute("""
        SELECT COUNT(*) AS n_lotes, COALESCE(SUM(qtd_atual),0) AS total
        FROM embriao_lote WHERE qtd_atual>0
    """).fetchone()
    estoque_t = db.execute(
        "SELECT COUNT(*) AS n FROM estoque_touros WHERE vendido=0"
    ).fetchone()
    return {
        "rodada": rodada['nome'],
        "matrizes_ativas": kpis['total_matrizes'],
        "iciagen_medio": round(kpis['iciagen_avg'] or 0, 2),
        "idesm_medio": round(kpis['idesm_avg'] or 0, 2),
        "rmat_medio": round(kpis['rmat_avg'] or 0, 2),
        "ipp_medio_meses": round(kpis['ipp_avg'] or 0, 1) if kpis['ipp_avg'] else None,
        "iep_medio_meses": round(kpis['iep_avg'] or 0, 1) if kpis['iep_avg'] else None,
        "matrizes_com_ceip": kpis['ceip_total'] or 0,
        "matrizes_genotipadas": kpis['genotipadas'] or 0,
        "matrizes_precoces": kpis['precoces'] or 0,
        "distribuicao_categoria": [dict(r) for r in cats],
        "estoque_embrioes": {"lotes_com_saldo": emb['n_lotes'], "total_disponivel": emb['total']},
        "estoque_touros_venda_disponiveis": estoque_t['n']
    }


def _pecia_execute_tool(db, name, args):
    args = args or {}
    try:
        if name == "consultar_animal":
            return _pecia_tool_consultar_animal(db, args.get('brinco'))
        if name == "genealogia":
            return _pecia_tool_genealogia(db, args.get('brinco'))
        if name == "top_animais_por_indice":
            return _pecia_tool_top_indice(
                db,
                args.get('indice'),
                args.get('categoria') or 'TODAS',
                args.get('limit') or 10,
                args.get('ordem') or 'melhores'
            )
        if name == "estoque_embrioes":
            return _pecia_tool_estoque_embrioes(
                db,
                args.get('doadora'),
                args.get('touro'),
                args.get('apenas_disponiveis', True)
            )
        if name == "estoque_touros_venda":
            return _pecia_tool_estoque_touros_venda(
                db,
                args.get('grupo'),
                args.get('apenas_disponiveis', True)
            )
        if name == "resumo_rebanho":
            return _pecia_tool_resumo_rebanho(db)
        if name == "gerar_relatorio_pdf":
            tipo = args.get('tipo')
            apenas_disp = args.get('apenas_disponiveis', True)
            logo_path = os.path.join(app.root_path, 'static', 'logo.jpg')
            if tipo == 'estoque_touros':
                return gerar_pdf_estoque_touros(
                    db, REPORTS_DIR, logo_path,
                    grupo=args.get('filtro_grupo'),
                    apenas_disponiveis=apenas_disp,
                )
            if tipo == 'estoque_embrioes':
                return gerar_pdf_estoque_embrioes(
                    db, REPORTS_DIR, logo_path,
                    doadora=args.get('filtro_doadora'),
                    touro=args.get('filtro_touro'),
                    apenas_disponiveis=apenas_disp,
                )
            return {"erro": f"Tipo de relatório não suportado: {tipo}"}
        return {"erro": f"Ferramenta desconhecida: {name}"}
    except Exception as e:
        return {"erro": f"Falha na ferramenta {name}: {str(e)}"}


@app.route('/relatorios/<path:filename>')
@login_required
def serve_relatorio(filename):
    """Serve PDFs gerados pela PecIA, autenticado."""
    return send_from_directory(
        REPORTS_DIR, filename,
        as_attachment=False, mimetype='application/pdf'
    )


@app.route('/api/pecia/chat', methods=['POST'])
@api_login_required
def api_pecia_chat():
    """Chat conversacional com Claude + tool use sobre o banco do rebanho (read-only)."""
    body = request.get_json(silent=True) or {}
    user_messages = body.get('messages') or []
    if not isinstance(user_messages, list) or not user_messages:
        return jsonify({'erro': 'Faltam messages no body'}), 400

    convo = []
    for m in user_messages:
        role = m.get('role')
        content = m.get('content')
        if role not in ('user', 'assistant') or not isinstance(content, str):
            continue
        convo.append({"role": role, "content": content})
    if not convo or convo[0]['role'] != 'user':
        return jsonify({'erro': 'Primeira mensagem precisa ser do usuário'}), 400

    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'erro': 'API key não configurada. Adicione ANTHROPIC_API_KEY ao .env'}), 500

    db = get_db()
    client = anthropic.Anthropic(api_key=api_key)

    tools_with_cache = [dict(t) for t in PECIA_TOOLS]
    tools_with_cache[-1] = {**tools_with_cache[-1], "cache_control": {"type": "ephemeral"}}

    messages = list(convo)
    attachments = []
    MAX_LOOPS = 6
    try:
        for _ in range(MAX_LOOPS):
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=4096,
                system=PECIA_SYSTEM_PROMPT,
                tools=tools_with_cache,
                messages=messages,
            )

            if response.stop_reason == "tool_use":
                messages.append({"role": "assistant", "content": response.content})
                tool_results = []
                for block in response.content:
                    if block.type == "tool_use":
                        result = _pecia_execute_tool(db, block.name, block.input)
                        if isinstance(result, dict) and result.get('pdf_url'):
                            attachments.append({
                                'url': result['pdf_url'],
                                'filename': result.get('filename', 'relatorio.pdf'),
                                'size_kb': result.get('size_kb', 0),
                                'tipo': result.get('tipo', ''),
                                'total_itens': result.get('total_itens', 0),
                            })
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": json.dumps(result, ensure_ascii=False, default=str)
                        })
                messages.append({"role": "user", "content": tool_results})
                continue

            final_text = next(
                (b.text for b in response.content if b.type == "text"),
                ""
            )
            return jsonify({
                "reply": final_text or "(sem resposta)",
                "stop_reason": response.stop_reason,
                "attachments": attachments,
                "usage": {
                    "input_tokens": response.usage.input_tokens,
                    "output_tokens": response.usage.output_tokens,
                    "cache_read_input_tokens": getattr(response.usage, 'cache_read_input_tokens', 0),
                    "cache_creation_input_tokens": getattr(response.usage, 'cache_creation_input_tokens', 0),
                }
            })

        return jsonify({"erro": "Limite de iterações de ferramenta atingido"}), 500

    except anthropic.APIError as e:
        return jsonify({'erro': f'Erro na API Claude: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'erro': f'Erro inesperado: {str(e)}'}), 500


# ── Init & Run ─────────────────────────────────────────────────────

with app.app_context():
    init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5000)
